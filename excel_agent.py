"""
🏢 ENTERPRISE EXCEL AUTOMATION AI AGENT
========================================

An enterprise-grade autonomous agent that intelligently analyzes, modifies, 
and enhances Excel files based on business language queries.

You are NOT a chatbot. You are a decision-making Excel agent.

ARCHITECTURE:
-------------
1. Excel Structure Analyzer - Deep file structure analysis
2. Business Query Parser - Converts business language to operations
3. Decision Engine - Enterprise-grade planning and validation
4. Execution Engine - Safe, auditable modifications
5. Validation & Quality Control - Data integrity enforcement
6. Reporting Engine - Clear explanations and change logs

SUPPORTED OPERATIONS:
--------------------
1. Calculations (Sum, Avg, Min, Max, %, etc.)
2. Data Transformation (Add/Rename/Remove columns, Sort, Filter)
3. Data Validation & Quality (Dropdowns, Error detection)
4. Conditional Formatting (Color coding, KPIs, Data bars)
5. Aggregation & Analytics (Pivot tables, Group by)
6. Lookup & Reference (XLOOKUP, INDEX-MATCH, Joins)
7. Visualization (Charts: Bar, Pie, Line, KPI dashboards)
8. Reporting & Governance (Summaries, Dashboards, Protection)
"""

from groq import Groq
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
import json
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
import os
import re


# ==========================================
# SYSTEM PROMPT - AGENT IDENTITY
# ==========================================
SYSTEM_PROMPT = """You are an Enterprise Excel Automation AI Agent, not a chatbot.

IDENTITY:
- You are a professional decision-making agent specialized in Excel operations
- You operate with enterprise-grade precision and governance
- You prioritize data integrity, auditability, and clear communication

CORE RESPONSIBILITIES:
1. Analyze Excel file structures with deep understanding
2. Parse business language queries into technical operations
3. Create risk-assessed decision plans before execution
4. Execute operations safely with complete audit trails
5. Provide clear, professional explanations of all changes

OPERATIONAL PRINCIPLES:
- Always create new files (never overwrite originals)
- Log every operation for audit compliance
- Assess risk before execution (low/medium/high)
- Request confirmation for high-risk operations
- Use exact column names from the actual file
- Provide before/after previews when possible

TONE & STYLE:
- Professional and concise
- Use business language, not technical jargon
- Provide structured, actionable responses
- Show confidence in recommendations
- Admit when clarification is needed

YOU ARE NOT:
- A general chatbot for casual conversation
- A data science consultant
- A replacement for human judgment on critical decisions
"""


# ==========================================
# DECISION DATA STRUCTURE
# ==========================================
@dataclass
class AgentDecision:
    """Enterprise-grade structured decision object before Excel modification"""
    operation_type: str  # calculation | transformation | validation | formatting | aggregation | lookup | visualization | reporting
    sub_operation: str  # Specific operation like "sum", "pivot", "xlookup", "bar_chart"
    source_columns: List[str]
    target_columns: Optional[List[str]]  # Can be multiple for complex operations
    sheet_name: str
    execution_scope: str  # all_rows | filtered_rows | specific_range
    risk_level: str  # low | medium | high
    assumptions: List[str]
    requires_confirmation: bool  # For high-risk operations
    change_description: str  # Plain English explanation
    formula_logic: str = ""  # Formula or logic description for the operation
    is_read_only: bool = False  # True for display/view operations that don't modify data
    
    def to_dict(self) -> Dict:
        return {
            "operation_type": self.operation_type,
            "sub_operation": self.sub_operation,
            "source_columns": self.source_columns,
            "target_columns": self.target_columns,
            "sheet_name": self.sheet_name,
            "execution_scope": self.execution_scope,
            "risk_level": self.risk_level,
            "assumptions": self.assumptions,
            "requires_confirmation": self.requires_confirmation,
            "change_description": self.change_description,
            "formula_logic": self.formula_logic,
            "is_read_only": self.is_read_only
        }


# ==========================================
# AGENT TOOLS (Excel Operations)
# ==========================================
EXCEL_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "analyze_excel_structure",
            "description": "Analyzes the Excel file structure including sheets, columns, data types, and sample data",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "Specific sheet to analyze (optional, analyzes first sheet if not provided)"
                    }
                },
                "required": ["file_path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "parse_user_query",
            "description": "Parses natural language query to extract operation type, columns involved, and intended action",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "User's natural language query"
                    },
                    "available_columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of available column names in the Excel file"
                    }
                },
                "required": ["query", "available_columns"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_decision_plan",
            "description": "Creates a structured decision plan for Excel modification based on parsed query and Excel structure",
            "parameters": {
                "type": "object",
                "properties": {
                    "operation_type": {
                        "type": "string",
                        "enum": ["calculation", "transformation", "validation", "formatting", "aggregation", "lookup", "visualization", "reporting"],
                        "description": "Type of Excel operation"
                    },
                    "source_columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Source columns involved in the operation"
                    },
                    "target_columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Names of new columns to create or existing columns to modify"
                    },
                    "sub_operation": {
                        "type": "string",
                        "description": "Specific operation: sum, average, pivot, xlookup, bar_chart, etc."
                    },
                    "execution_scope": {
                        "type": "string",
                        "enum": ["all_rows", "filtered_rows", "specific_range"],
                        "description": "Which rows to apply the operation to"
                    }
                },
                "required": ["operation_type", "source_columns", "sub_operation"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "execute_excel_operation",
            "description": "Executes the planned Excel operation safely on the file",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Path to the Excel file"
                    },
                    "decision_plan": {
                        "type": "object",
                        "description": "The approved decision plan from create_decision_plan"
                    },
                    "output_path": {
                        "type": "string",
                        "description": "Path to save the modified Excel file"
                    }
                },
                "required": ["file_path", "decision_plan", "output_path"]
            }
        }
    }
]


# ==========================================
# EXCEL AUTOMATION AGENT CLASS
# ==========================================
class ExcelAgent:
    """
    Enterprise-grade Excel Automation AI Agent
    
    Decision-making agent for business Excel operations.
    NOT a chatbot - a professional automation tool.
    """
    
    def __init__(self, api_key: str):
        """Initialize the Enterprise Excel Agent with Groq API"""
        self.client = Groq(api_key=api_key)
        self.model = "llama-3.3-70b-versatile"
        self.system_prompt = SYSTEM_PROMPT
        self.conversation_history = []
        self.change_log = []  # Track all changes for audit
    
    def _make_json_safe(self, data: Any) -> Any:
        """Convert Timestamps and other non-JSON-serializable types to strings"""
        if isinstance(data, dict):
            return {k: self._make_json_safe(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._make_json_safe(item) for item in data]
        elif isinstance(data, pd.Timestamp):
            return data.strftime('%Y-%m-%d %H:%M:%S')
        elif hasattr(data, 'isoformat'):  # datetime objects
            return data.isoformat()
        elif pd.isna(data):
            return None
        else:
            return data
    
    def _detect_header_row(self, df: pd.DataFrame, max_rows_to_scan: int = 10) -> int:
        """
        Detect the true header row by finding the row with highest non-null ratio
        and most string-like values (headers are typically text).
        
        Returns:
            Index of the detected header row (0-based)
        """
        if len(df) == 0:
            return 0
        
        rows_to_check = min(max_rows_to_scan, len(df))
        best_row = 0
        best_score = 0
        
        for i in range(rows_to_check):
            row = df.iloc[i]
            # Calculate score based on:
            # 1. Non-null ratio
            # 2. String values (headers are usually text)
            # 3. Unique values (headers should be unique)
            
            non_null_count = row.notna().sum()
            non_null_ratio = non_null_count / len(row) if len(row) > 0 else 0
            
            # Count string values that look like headers (not numbers)
            string_count = 0
            for val in row.values:  # Use .values to get numpy array for safe iteration
                try:
                    is_valid = val is not None and pd.notna(val)
                except (ValueError, TypeError):
                    is_valid = val is not None
                
                if is_valid:
                    val_str = str(val).strip()
                    # Check if it looks like a header (text, not pure number)
                    try:
                        float(val_str.replace(',', ''))
                        # It's a number, less likely to be header
                    except ValueError:
                        string_count += 1
            
            string_ratio = string_count / len(row) if len(row) > 0 else 0
            
            # Check uniqueness
            unique_ratio = row.nunique() / len(row) if len(row) > 0 else 0
            
            # Combined score
            score = (non_null_ratio * 0.4) + (string_ratio * 0.4) + (unique_ratio * 0.2)
            
            if score > best_score:
                best_score = score
                best_row = i
        
        return best_row
    
    def _normalize_header(self, header: Any) -> str:
        """
        Normalize a single header value:
        - Remove line breaks
        - Trim whitespace
        - Handle None/NaN
        - Handle Series (take first value)
        """
        # Handle Series - take first value if passed
        if isinstance(header, pd.Series):
            header = header.iloc[0] if len(header) > 0 else None
        
        # Handle None/NaN - use try/except to avoid Series ambiguity
        try:
            if header is None or pd.isna(header):
                return ""
        except (ValueError, TypeError):
            # pd.isna can fail on some types, treat as valid
            pass
        
        header_str = str(header)
        # Remove line breaks and extra whitespace
        header_str = header_str.replace('\n', ' ').replace('\r', ' ')
        header_str = ' '.join(header_str.split())  # Normalize whitespace
        return header_str.strip()
    
    def _normalize_headers(self, df: pd.DataFrame, header_row: int = 0) -> pd.DataFrame:
        """
        Normalize DataFrame headers:
        - Use detected header row as column names
        - Remove line breaks, trim spaces
        - Handle multi-row headers by flattening
        - Drop pre-header rows
        
        Returns:
            DataFrame with normalized headers
        """
        if header_row == 0:
            # Just normalize existing headers
            new_columns = []
            for i, col in enumerate(df.columns):
                normalized = self._normalize_header(col)
                if normalized == "" or normalized.startswith("Unnamed"):
                    # Try to get from first data row if header is empty
                    if len(df) > 0:
                        try:
                            # Use .iat for guaranteed scalar access
                            first_val = df.iat[0, i]
                            # Ensure first_val is a scalar, not a Series
                            if isinstance(first_val, pd.Series):
                                first_val = first_val.iloc[0] if len(first_val) > 0 else None
                            if pd.notna(first_val) and not str(first_val).replace('.','').replace('-','').isdigit():
                                normalized = self._normalize_header(first_val)
                        except (IndexError, ValueError):
                            pass  # Keep the default Column_N name
                if normalized == "":
                    normalized = f"Column_{i+1}"
                new_columns.append(normalized)
            df.columns = new_columns
            return df
        
        # Use the detected header row as new column names
        new_columns = []
        
        for i in range(len(df.columns)):
            try:
                # Use .iat for guaranteed scalar access
                val = df.iat[header_row, i]
                # Ensure val is a scalar, not a Series
                if isinstance(val, pd.Series):
                    val = val.iloc[0] if len(val) > 0 else None
            except (IndexError, ValueError):
                val = None
            normalized = self._normalize_header(val)
            if normalized == "":
                normalized = f"Column_{i+1}"
            new_columns.append(normalized)
        
        # Create new DataFrame without pre-header rows
        new_df = df.iloc[header_row + 1:].copy()
        new_df.columns = new_columns
        new_df = new_df.reset_index(drop=True)
        
        return new_df
    
    def _classify_columns(self, df: pd.DataFrame) -> Dict[str, List[str]]:
        """
        Classify columns semantically:
        - Entity columns: Text columns that identify who/what (names, IDs)
        - Measure columns: Numeric columns for calculations/visualization
        - Context columns: Categorical/grouping columns
        - Date columns: Temporal data
        
        Returns:
            Dictionary with classified column lists
        """
        classification = {
            "entity_columns": [],      # Who/what the data is about (names, IDs)
            "measure_columns": [],     # Numeric values for calculation
            "context_columns": [],     # Categories, groups, labels
            "date_columns": [],        # Date/time columns
            "unknown_columns": []      # Cannot classify
        }
        
        # Use column index to avoid issues with duplicate column names
        for col_idx, col in enumerate(df.columns):
            try:
                # Use iloc with column index to avoid duplicate column name issues
                col_data = df.iloc[:, col_idx].dropna()
                if len(col_data) == 0:
                    classification["unknown_columns"].append(col)
                    continue
                
                # Check if date column
                if pd.api.types.is_datetime64_any_dtype(df.iloc[:, col_idx]):
                    classification["date_columns"].append(col)
                    continue
                
                # Check if numeric
                if pd.api.types.is_numeric_dtype(df.iloc[:, col_idx]):
                    classification["measure_columns"].append(col)
                    continue
                
                # For text columns, classify as entity or context
                col_lower = str(col).lower()  # Ensure col is string
                unique_ratio = df.iloc[:, col_idx].nunique() / len(df) if len(df) > 0 else 0
            except Exception:
                classification["unknown_columns"].append(col)
                continue
            
            # Entity indicators: high uniqueness, contains name/id keywords
            entity_keywords = ['name', 'id', 'student', 'employee', 'person', 'customer', 'user', 'member']
            is_entity = any(kw in col_lower for kw in entity_keywords) or unique_ratio > 0.5
            
            # Context indicators: low uniqueness, contains category keywords
            context_keywords = ['type', 'category', 'group', 'class', 'grade', 'level', 'status', 'segment', 'division', 'section']
            is_context = any(kw in col_lower for kw in context_keywords) or unique_ratio < 0.3
            
            if is_entity and not is_context:
                classification["entity_columns"].append(col)
            elif is_context:
                classification["context_columns"].append(col)
            else:
                # Default: if unique ratio is high, it's entity; otherwise context
                if unique_ratio > 0.5:
                    classification["entity_columns"].append(col)
                else:
                    classification["context_columns"].append(col)
        
        return classification
    
    def _normalize_cell_value(self, value: Any) -> str:
        """
        Normalize a cell value for matching:
        - Remove line breaks
        - Trim whitespace
        - Convert to string
        - Handle Series (take first value)
        """
        # Handle Series - take first value if passed
        if isinstance(value, pd.Series):
            value = value.iloc[0] if len(value) > 0 else None
        
        # Handle None/NaN with try/except to avoid Series ambiguity
        try:
            if value is None or pd.isna(value):
                return ""
        except (ValueError, TypeError):
            pass
        
        val_str = str(value)
        # Remove line breaks
        val_str = val_str.replace('\n', ' ').replace('\r', ' ')
        # Normalize whitespace
        val_str = ' '.join(val_str.split())
        return val_str.strip()
    
    def _token_match(self, search_term: str, value: str) -> bool:
        """
        Token-based matching for flexible entity/group matching.
        Matches if search term tokens are found in value tokens.
        
        Example: "Chandra" matches "MYP 1 CHANDRA (Ms. Deepa)"
        Example: "maths" matches "Mathematics"
        """
        if not search_term or not value:
            return False
        
        # Normalize both
        search_normalized = self._normalize_cell_value(search_term).lower()
        value_normalized = self._normalize_cell_value(value).lower()
        
        # Direct contains check (either direction)
        if search_normalized in value_normalized:
            return True
        if value_normalized in search_normalized:
            return True
        
        # Partial match check (for cases like "maths" in "mathematics")
        # Check if search term is a prefix/substring of any word in value
        value_words = value_normalized.replace('(', ' ').replace(')', ' ').split()
        for word in value_words:
            if search_normalized in word or word in search_normalized:
                return True
            # Check prefix match (at least 3 chars)
            if len(search_normalized) >= 3 and word.startswith(search_normalized[:3]):
                return True
            if len(word) >= 3 and search_normalized.startswith(word[:3]):
                return True
        
        # Token-based check
        search_tokens = set(search_normalized.split())
        value_tokens = set(value_normalized.split())
        
        # Match if any search token is found in value tokens
        for token in search_tokens:
            if any(token in vt for vt in value_tokens):
                return True
        
        return False
    
    def _extract_query_intent(self, query: str, column_classification: Dict) -> Dict:
        """
        Extract structured intent from user query.
        
        Returns:
            Dictionary with intent, target entities, filters, measures, output type
        """
        query_lower = query.lower()
        
        intent = {
            "operation": "unknown",
            "target_entities": [],
            "filters": [],
            "measures": [],
            "output_type": "data",
            "aggregation": None
        }
        
        # Detect operation type
        if any(word in query_lower for word in ['chart', 'pie', 'bar', 'line', 'graph', 'plot', 'visualize']):
            intent["operation"] = "visualization"
            intent["output_type"] = "chart"
            if 'pie' in query_lower:
                intent["chart_type"] = "pie"
            elif 'line' in query_lower:
                intent["chart_type"] = "line"
            else:
                intent["chart_type"] = "bar"
        elif any(word in query_lower for word in ['filter', 'where', 'only', 'show me']):
            intent["operation"] = "filter"
        elif any(word in query_lower for word in ['sort', 'order', 'rank']):
            intent["operation"] = "sort"
        elif any(word in query_lower for word in ['sum', 'total', 'average', 'mean', 'count', 'max', 'min']):
            intent["operation"] = "aggregation"
            if 'sum' in query_lower or 'total' in query_lower:
                intent["aggregation"] = "sum"
            elif 'average' in query_lower or 'mean' in query_lower:
                intent["aggregation"] = "mean"
            elif 'count' in query_lower:
                intent["aggregation"] = "count"
        elif any(word in query_lower for word in ['print', 'show', 'display', 'view', 'list']):
            intent["operation"] = "display"
            intent["output_type"] = "data"
        else:
            intent["operation"] = "transform"
        
        # Extract potential entity references from query
        # Match against known entity columns
        for entity_col in column_classification.get("entity_columns", []):
            if entity_col.lower() in query_lower:
                intent["target_entities"].append(entity_col)
        
        # Extract potential measure references
        for measure_col in column_classification.get("measure_columns", []):
            if measure_col.lower() in query_lower:
                intent["measures"].append(measure_col)
        
        return intent

    def _generate_pandas_code(self, query: str, excel_structure: Dict) -> Dict:
        """
        Generate executable pandas code using LLM for any Excel operation.
        This is the core of the intelligent agent - LLM writes the actual code.
        
        Returns:
            Dictionary with generated code, explanation, and metadata
        """
        columns = excel_structure.get("columns", [])
        numeric_cols = excel_structure.get("numeric_columns", [])
        text_cols = excel_structure.get("text_columns", [])
        sample_data = excel_structure.get("sample_data", [])
        total_rows = excel_structure.get("total_rows", 0)
        
        # Get semantic column classification
        entity_cols = excel_structure.get("entity_columns", [])
        measure_cols = excel_structure.get("measure_columns", [])
        context_cols = excel_structure.get("context_columns", [])
        date_cols = excel_structure.get("date_columns", [])
        
        code_gen_prompt = f"""You are an Enterprise Excel Automation Agent - an expert Python/Pandas code generator.

AVAILABLE DATA:
- DataFrame variable: `df` (already loaded, do NOT reload)
- All Columns: {columns}
- Total rows: {total_rows}

COLUMN CLASSIFICATION (Use this to understand the data):
- Entity Columns (who/what): {entity_cols if entity_cols else 'None detected'}
- Measure Columns (numeric values): {measure_cols if measure_cols else numeric_cols}
- Context/Grouping Columns: {context_cols if context_cols else 'None detected'}
- Date Columns: {date_cols if date_cols else 'None detected'}

SAMPLE DATA (CRITICAL - Study this to understand actual values):
{self._make_json_safe(sample_data[:5])}

USER REQUEST: "{query}"

YOUR TASK:
Generate Python pandas code to fulfill the user's request. You MUST handle ANY Excel structure.

MANDATORY RULES:
1. DataFrame is `df` - do NOT use pd.read_excel()
2. Final result MUST be stored in `df`
3. Use EXACT column names from the list (case-sensitive)
4. matplotlib.pyplot is available as `plt` for charts

DATA NORMALIZATION (ALWAYS DO THIS):
5. Normalize cell values before matching:
   - Remove line breaks: df['col'] = df['col'].astype(str).str.replace(r'\\n|\\r', ' ', regex=True)
   - Strip whitespace: df['col'] = df['col'].str.strip()
6. Convert to numeric when needed:
   df['col'] = pd.to_numeric(df['col'], errors='coerce')
7. Drop NaN only when necessary: df = df.dropna(subset=['col'])

FUZZY/TOKEN MATCHING (CRITICAL):
8. ALWAYS use case-insensitive partial matching:
   df[df['col'].astype(str).str.contains('search_term', case=False, na=False)]
9. User says "Chandra" → match "MYP 1 CHANDRA", "Chandra Section", etc.
10. User says "maths" → match "Mathematics", "Math", "MATH", etc.
11. For multi-word search, match ANY token:
    df[df['col'].astype(str).str.contains('word1|word2', case=False, na=False)]

VISUALIZATION:
12. matplotlib is pre-loaded as `plt`
13. For pie: plt.figure(figsize=(10,8)); plt.pie(values, labels=labels, autopct='%1.1f%%')
14. For bar: plt.figure(figsize=(12,6)); plt.bar(x, y)
15. ALWAYS: plt.tight_layout(); plt.savefig('chart.png', dpi=150, bbox_inches='tight'); plt.close()
16. Filter data BEFORE plotting - don't plot all rows

INTELLIGENT ASSUMPTIONS:
17. NEVER ask for clarification - make the most reasonable assumption
18. If user mentions a term, find it in ANY column using fuzzy match
19. If ambiguous, pick the first/most logical match and proceed
20. Study sample data to understand what values actually exist

RESPONSE FORMAT (JSON only):
{{
    "pandas_code": "# Normalize and process\\ndf['col'] = df['col'].astype(str).str.strip()\\n...",
    "operation_description": "Brief description of what the code does",
    "columns_affected": ["list", "of", "columns"],
    "is_filter_operation": true/false,
    "is_read_only": true/false,
    "risk_level": "low/medium/high",
    "expected_row_change": "same/increase/decrease/unknown"
}}

Return ONLY valid JSON, no markdown or explanation."""

        messages = [
            {"role": "system", "content": "You are a pandas code generator. Return ONLY valid JSON with executable Python code."},
            {"role": "user", "content": code_gen_prompt}
        ]
        
        response = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.1,
            max_tokens=1000
        )
        
        return self._extract_json_from_response(response)
    
    def _execute_generated_code(self, df: pd.DataFrame, code: str, query: str) -> tuple:
        """
        Safely execute LLM-generated pandas code.
        
        Returns:
            Tuple of (modified_df, success_bool, error_message)
        """
        # Import safe visualization libraries
        try:
            import matplotlib
            matplotlib.use('Agg')  # Non-interactive backend for server use
            import matplotlib.pyplot as plt
            has_matplotlib = True
        except ImportError:
            has_matplotlib = False
            plt = None
        
        # Create a safe execution environment with pre-loaded safe libraries
        safe_globals = {
            'pd': pd,
            'np': __import__('numpy'),
            're': re,
            'df': df.copy(),  # Work on a copy for safety
            'datetime': __import__('datetime'),
        }
        
        # Add matplotlib if available (for visualizations)
        if has_matplotlib:
            safe_globals['plt'] = plt
            safe_globals['matplotlib'] = matplotlib
        
        # Add common pandas functions
        safe_globals['DataFrame'] = pd.DataFrame
        safe_globals['Series'] = pd.Series
        
        try:
            # Clean the code - remove import statements (we pre-load safe ones)
            clean_code = code.strip()
            
            # Remove import lines since we pre-load safe libraries
            code_lines = clean_code.split('\n')
            filtered_lines = []
            for line in code_lines:
                line_stripped = line.strip().lower()
                # Allow matplotlib/numpy imports by skipping them (already loaded)
                if line_stripped.startswith('import matplotlib') or \
                   line_stripped.startswith('import numpy') or \
                   line_stripped.startswith('from matplotlib') or \
                   line_stripped.startswith('import datetime'):
                    continue  # Skip - already pre-loaded
                filtered_lines.append(line)
            clean_code = '\n'.join(filtered_lines)
            
            # Remove any dangerous operations
            dangerous_patterns = [
                'exec(', 'eval(', 'open(', 'os.', 'subprocess',
                'shutil', '__', 'globals(', 'locals(', 'compile(',
                'read_excel', 'to_excel', 'read_csv', 'to_csv',
                'import os', 'import sys', 'import subprocess'
            ]
            
            for pattern in dangerous_patterns:
                if pattern in clean_code.lower():
                    return df, False, f"Security: '{pattern}' is not allowed in generated code"
            
            # Execute the code
            exec(clean_code, safe_globals)
            
            # Get the modified DataFrame
            result_df = safe_globals.get('df', df)
            
            if not isinstance(result_df, pd.DataFrame):
                return df, False, "Code did not produce a valid DataFrame"
            
            return result_df, True, None
            
        except Exception as e:
            return df, False, f"Code execution error: {str(e)}"
    
    def process_query_with_code_generation(self, query: str, file_path: str, 
                                            sheet_name: Optional[str] = None) -> Dict:
        """
        Process user query using LLM code generation approach.
        This is the intelligent agent that lets LLM handle any Excel operation.
        
        Returns:
            Dictionary with execution results
        """
        # Step 1: Analyze Excel structure
        excel_structure = self.analyze_excel_structure(file_path, sheet_name)
        if "error" in excel_structure:
            return {"status": "failed", "error": excel_structure["error"]}
        
        target_sheet = excel_structure.get("analyzed_sheet", "Sheet1")
        
        # Step 2: Generate pandas code using LLM
        print(f"\n   🤖 Generating pandas code for: '{query}'")
        code_response = self._generate_pandas_code(query, excel_structure)
        
        if not code_response or "pandas_code" not in code_response:
            return {"status": "failed", "error": "Failed to generate pandas code"}
        
        pandas_code = code_response.get("pandas_code", "")
        operation_desc = code_response.get("operation_description", "Execute operation")
        is_read_only = code_response.get("is_read_only", False)
        risk_level = code_response.get("risk_level", "medium")
        
        print(f"   📝 Generated Code:\n      {pandas_code.replace(chr(10), chr(10) + '      ')}")
        print(f"   📋 Operation: {operation_desc}")
        print(f"   ⚠️  Risk Level: {risk_level.upper()}")
        
        # Step 3: Load DataFrame
        df = pd.read_excel(file_path, sheet_name=target_sheet)
        original_row_count = len(df)
        
        # Step 4: Execute the generated code
        print(f"\n   ⚙️  Executing generated code...")
        result_df, success, error = self._execute_generated_code(df, pandas_code, query)
        
        if not success:
            return {"status": "failed", "error": error}
        
        new_row_count = len(result_df)
        
        # Step 5: Save results (if not read-only)
        if is_read_only:
            return {
                "status": "success",
                "message": operation_desc,
                "rows_affected": new_row_count,
                "output_file": None,
                "is_read_only": True,
                "data_preview": self._make_json_safe(result_df.head(10).to_dict('records')),
                "generated_code": pandas_code
            }
        
        # Generate output filename
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        dir_name = os.path.dirname(file_path)
        counter = 1
        while True:
            output_path = os.path.join(dir_name, f"{base_name}_modified_{counter}.xlsx")
            if not os.path.exists(output_path):
                break
            counter += 1
        
        # Save with all sheets preserved
        self._save_with_all_sheets(file_path, result_df, target_sheet, output_path)
        
        return {
            "status": "success",
            "message": operation_desc,
            "rows_before": original_row_count,
            "rows_after": new_row_count,
            "rows_affected": abs(original_row_count - new_row_count) if original_row_count != new_row_count else new_row_count,
            "output_file": output_path,
            "is_read_only": False,
            "data_preview": self._make_json_safe(result_df.head(10).to_dict('records')),
            "generated_code": pandas_code,
            "columns_affected": code_response.get("columns_affected", [])
        }
        
    def analyze_excel_structure(self, file_path: str, sheet_name: Optional[str] = None, auto_detect_headers: bool = True) -> Dict:
        """
        Analyzes Excel file structure with intelligent header detection and normalization.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Optional specific sheet name
            auto_detect_headers: If True, automatically detect and normalize headers
        
        Returns:
            Dictionary containing sheets, columns, data types, sample data, and column classification
        """
        try:
            # Load Excel file
            xl_file = pd.ExcelFile(file_path)
            sheets = xl_file.sheet_names
            
            # Use first sheet if not specified
            target_sheet = sheet_name if sheet_name else sheets[0]
            
            # Read the sheet (header=None to get raw data first for header detection)
            df_raw = pd.read_excel(file_path, sheet_name=target_sheet, header=None)
            
            # Detect and normalize headers if enabled
            header_row = 0
            header_detection_info = None
            
            if auto_detect_headers and len(df_raw) > 0:
                # Check if we have "Unnamed" style columns or need detection
                df_test = pd.read_excel(file_path, sheet_name=target_sheet)
                unnamed_count = sum(1 for col in df_test.columns if str(col).startswith('Unnamed:'))
                
                if unnamed_count > len(df_test.columns) * 0.3:  # More than 30% unnamed
                    # Detect the true header row
                    header_row = self._detect_header_row(df_raw)
                    header_detection_info = f"Auto-detected header at row {header_row + 1}"
                    
                    # Normalize headers
                    df = self._normalize_headers(df_raw, header_row)
                else:
                    # Use standard reading but normalize existing headers
                    df = df_test
                    df = self._normalize_headers(df, header_row=0)
            else:
                df = pd.read_excel(file_path, sheet_name=target_sheet)
            
            # Check for any remaining problematic column names
            unnamed_columns = [col for col in df.columns if str(col).startswith('Unnamed:') or str(col).startswith('Column_')]
            has_header_issues = len(unnamed_columns) > 0
            header_warning = None
            
            if has_header_issues:
                header_warning = f"⚠️  {len(unnamed_columns)} columns could not be auto-named. Check the source file structure."
            
            # Classify columns semantically
            column_classification = self._classify_columns(df)
            
            # Analyze structure
            analysis = {
                "file_name": os.path.basename(file_path),
                "available_sheets": sheets,
                "analyzed_sheet": target_sheet,
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": list(df.columns),
                "data_types": {col: str(dtype) for col, dtype in df.dtypes.items()},
                "sample_data": self._make_json_safe(df.head(5).to_dict('records')),
                "null_counts": self._make_json_safe(df.isnull().sum().to_dict()),
                "numeric_columns": df.select_dtypes(include=['number']).columns.tolist(),
                "text_columns": df.select_dtypes(include=['object']).columns.tolist(),
                "has_header_issues": has_header_issues,
                "unnamed_columns_count": len(unnamed_columns),
                "header_warning": header_warning,
                "header_detection_info": header_detection_info,
                "column_classification": column_classification,
                "entity_columns": column_classification.get("entity_columns", []),
                "measure_columns": column_classification.get("measure_columns", []),
                "context_columns": column_classification.get("context_columns", []),
                "date_columns": column_classification.get("date_columns", [])
            }
            
            return analysis
            
        except Exception as e:
            return {"error": str(e), "status": "failed"}
    
    def parse_user_query(self, query: str, excel_structure: Dict) -> Dict:
        """
        Enhanced AI parsing with multi-step reasoning and comprehensive prompt engineering
        
        Returns:
            Parsed query with operation type, columns, and intent
        """
        prompt = self._build_enhanced_parsing_prompt(query, excel_structure)
        
        # Use system prompt with enhanced user prompt
        messages = [
            {"role": "system", "content": self.system_prompt},
            {"role": "user", "content": prompt}
        ]
        
        response = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.1,  # Low temperature for precise parsing
            max_tokens=1500,  # Increased for detailed reasoning
            top_p=0.95
        )
        
        return self._extract_json_from_response(response)
    
    def _build_enhanced_parsing_prompt(self, query: str, excel_structure: Dict) -> str:
        """
        Build comprehensive prompt with multi-step reasoning for query parsing
        """
        available_columns = excel_structure.get("columns", [])
        numeric_columns = excel_structure.get("numeric_columns", [])
        text_columns = excel_structure.get("text_columns", [])
        sample_data = excel_structure.get("sample_data", [])
        total_rows = excel_structure.get('total_rows', 0)
        total_columns = excel_structure.get('total_columns', 0)
        available_sheets = excel_structure.get('available_sheets', [])
        current_sheet = excel_structure.get('analyzed_sheet', 'Sheet1')
        
        # Safely convert columns to strings
        columns_str = ', '.join([str(c) for c in available_columns])
        numeric_str = ', '.join([str(c) for c in numeric_columns]) if numeric_columns else 'None'
        text_str = ', '.join([str(c) for c in text_columns]) if text_columns else 'None'
        sheets_str = ', '.join(available_sheets) if available_sheets else current_sheet
        
        prompt = f"""You are an Enterprise Excel Operation Parser with deep domain expertise.

CONTEXT - EXCEL FILE ANALYSIS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 Total Rows: {total_rows}
📋 Total Columns: {total_columns}
📑 Available Sheets: {sheets_str}
📄 Currently Working On Sheet: {current_sheet}

AVAILABLE COLUMNS:
{columns_str}

NUMERIC COLUMNS (can be calculated):
{numeric_str}

TEXT COLUMNS (categories/labels):
{text_str}

SAMPLE DATA (First 3 rows for context):
{json.dumps(sample_data[:3], indent=2)}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

USER'S BUSINESS QUERY:
"{query}"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

YOUR TASK - MULTI-STEP REASONING:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

STEP 1: INTENT ANALYSIS
- What is the user trying to accomplish?
- Is this a calculation, visualization, transformation, or analysis?
- What business outcome do they expect?

STEP 2: COLUMN MAPPING
- Which columns are mentioned (explicitly or implicitly)?
- Map business terms to exact column names
- Identify if columns exist or need to be created

STEP 3: OPERATION CLASSIFICATION
- Determine the operation type and sub-operation
- Consider data types (numeric vs text)
- Assess technical feasibility

STEP 4: RISK ASSESSMENT
- Does this modify existing data? (higher risk)
- Does this delete information? (high risk)
- Is this reversible? (lower risk)

STEP 5: VALIDATION
- Are all required columns available?
- Is the operation technically possible?
- Any ambiguities that need clarification?

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

OUTPUT FORMAT - RETURN ONLY VALID JSON:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

TASK: Extract and return ONLY a valid JSON object (no markdown, no code blocks) with this exact structure:
{{
    "reasoning": {{
        "intent_summary": "Brief summary of what user wants",
        "column_mapping": {{"user_term": "actual_column_name"}},
        "operation_rationale": "Why this operation type was chosen",
        "risk_factors": ["factor1", "factor2"]
    }},
    "operation_type": "calculation | transformation | validation | formatting | aggregation | lookup | visualization | reporting",
    "sub_operation": "specific_operation_name",
    "source_columns": ["exact_column_name1", "exact_column_name2"],
    "needs_new_column": true,
    "suggested_column_names": ["New_Column_Name"],
    "operation_description": "Plain English description of what will happen",
    "formula_logic": "If calculation: 'sum(A, B)', if visualization: 'bar chart of Sales by Product'",
    "is_ambiguous": false,
    "clarification_needed": null,
    "risk_level": "low | medium | high",
    "execution_confidence": "high | medium | low",
    "alternative_interpretations": []
}}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

OPERATION TYPES & SUB-OPERATIONS CATALOG:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. CALCULATION
   • sum - Add multiple columns/values
   • average/mean - Calculate mean value
   • min/minimum - Find minimum value
   • max/maximum - Find maximum value
   • count - Count non-empty values
   • median - Calculate median
   • multiply - Multiply columns
   • divide - Divide columns
   • percentage - Calculate percentage (ratio or of total)
   • variance - Statistical variance
   • std_dev - Standard deviation

2. TRANSFORMATION
   • add_column - Create new empty column
   • rename_column - Rename existing column
   • remove_column - Delete column (HIGH RISK)
   • sort - Sort data by columns
   • filter - Filter rows by condition
   • deduplicate - Remove duplicate rows
   • split_column - Split column into multiple
   • merge_columns - Combine multiple columns
   • fill_missing - Fill null values
   • replace_values - Replace specific values

3. VALIDATION
   • dropdown - Create dropdown lists
   • range_check - Validate numeric ranges
   • error_detection - Find data quality issues
   • blank_check - Identify missing values
   • duplicate_check - Find duplicates
   • data_type_validation - Ensure correct data types

4. FORMATTING
   • conditional_color - Color cells by condition
   • highlight - Highlight specific values
   • data_bars - Add data bar visualization in cells
   • icon_sets - Add conditional icons
   • font_style - Change font properties
   • number_format - Format numbers (currency, %, etc.)

5. AGGREGATION
   • pivot_table - Create pivot table
   • group_by - Group and aggregate data
   • cross_tab - Cross-tabulation
   • summary_stats - Generate summary statistics
   • subtotals - Add subtotal rows

6. LOOKUP
   • xlookup - Modern lookup function
   • vlookup - Vertical lookup
   • index_match - INDEX-MATCH combination
   • join_sheets - Merge data from multiple sheets
   • fuzzy_match - Approximate matching

7. VISUALIZATION
   • bar_chart - Bar/column chart
   • pie_chart - Pie chart
   • line_chart - Line chart
   • scatter_plot - Scatter plot
   • kpi_dashboard - KPI dashboard
   • heatmap - Conditional formatting heatmap
   • sparklines - Mini charts in cells

8. REPORTING
   • summary_sheet - Create summary sheet
   • dashboard - Executive dashboard
   • change_log - Audit trail of changes
   • export_summary - Export summary report
   • data_profile - Statistical profile of data

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

BUSINESS LANGUAGE MAPPING (Natural to Technical):
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

CALCULATION TERMS:
"total", "sum up", "add together" → calculation + sum
"average", "mean", "typical" → calculation + average
"highest", "maximum", "largest" → calculation + max
"lowest", "minimum", "smallest" → calculation + min
"how many", "count" → calculation + count
"middle value", "median" → calculation + median
"ratio", "proportion" → calculation + percentage
"product", "times" → calculation + multiply
"split", "per" → calculation + divide

VISUALIZATION TERMS:
"chart", "graph", "plot", "visualize", "show me" → visualization
"bar chart", "column chart", "bars" → visualization + bar_chart
"pie chart", "pie", "distribution" → visualization + pie_chart
"line chart", "trend", "over time" → visualization + line_chart
"dashboard", "KPI", "metrics" → visualization + kpi_dashboard

TRANSFORMATION TERMS:
"sort", "order", "arrange" → transformation + sort
"remove duplicates", "unique values" → transformation + deduplicate
"split", "separate" → transformation + split_column
"combine", "merge", "concatenate" → transformation + merge_columns
"delete", "remove column" → transformation + remove_column (HIGH RISK)

AGGREGATION TERMS:
"pivot", "summarize by", "break down by" → aggregation + pivot_table
"group by", "aggregate" → aggregation + group_by
"subtotals", "totals for each" → aggregation + subtotals

FORMATTING TERMS:
"highlight", "color", "mark", "flag" → formatting + highlight
"format", "style" → formatting + conditional_color

LOOKUP TERMS:
"lookup", "match", "find", "search" → lookup + xlookup
"join", "merge sheets", "combine data" → lookup + join_sheets

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

CRITICAL RULES - MUST FOLLOW:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1. COLUMN NAME EXACTNESS
   • Use EXACT column names from "AVAILABLE COLUMNS" list
   • Do NOT invent or assume column names
   • If user says "column A" or "first column", map to actual name
   • If ambiguous, set is_ambiguous=true and request clarification

2. DATA TYPE AWARENESS
   • Only suggest calculations on numeric columns
   • Charts need: 1 text column (category) + 1+ numeric columns (values)
   • Don't try to sum text columns

3. RISK ASSESSMENT MATRIX
   • LOW RISK: Calculations, adding columns, visualizations, sorting
   • MEDIUM RISK: Formatting, renaming, filtering
   • HIGH RISK: Deleting columns, removing rows, cross-sheet operations

4. AMBIGUITY HANDLING
   • If multiple interpretations exist, list in alternative_interpretations
   • If critical info missing, set clarification_needed
   • If columns don't exist, suggest creating them

5. OUTPUT FORMAT
   • Return ONLY the JSON object
   • NO markdown code blocks (```json```)
   • NO explanatory text before/after JSON
   • Ensure all fields are present (use null if not applicable)

6. CHART-SPECIFIC LOGIC
   • For charts: identify category column (text) and value columns (numeric)
   • Set formula_logic to describe chart: "bar chart of [value] by [category]"
   • Example: "bar chart of Sales by Product"

7. CONFIDENCE SCORING
   • HIGH confidence: Clear query, all columns exist, single interpretation
   • MEDIUM confidence: Minor ambiguity, need assumption
   • LOW confidence: Missing columns, multiple interpretations possible

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

EXAMPLES OF CORRECT PARSING:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

EXAMPLE 1 - Calculation:
Query: "Calculate total revenue by summing Q1, Q2, Q3"
{{
    "reasoning": {{
        "intent_summary": "User wants to calculate total annual revenue from quarterly data",
        "column_mapping": {{"Q1": "Q1", "Q2": "Q2", "Q3": "Q3"}},
        "operation_rationale": "Sum operation on numeric columns to create annual total",
        "risk_factors": ["Creates new column - low risk"]
    }},
    "operation_type": "calculation",
    "sub_operation": "sum",
    "source_columns": ["Q1", "Q2", "Q3"],
    "needs_new_column": true,
    "suggested_column_names": ["Total_Revenue"],
    "operation_description": "Sum Q1, Q2, and Q3 to create Total_Revenue column",
    "formula_logic": "sum(Q1, Q2, Q3)",
    "is_ambiguous": false,
    "clarification_needed": null,
    "risk_level": "low",
    "execution_confidence": "high",
    "alternative_interpretations": []
}}

EXAMPLE 2 - Visualization:
Query: "Show me a bar chart of sales by product"
{{
    "reasoning": {{
        "intent_summary": "User wants visual representation of sales performance across products",
        "column_mapping": {{"sales": "Sales", "product": "Product"}},
        "operation_rationale": "Bar chart with Product as category and Sales as values",
        "risk_factors": ["Read-only operation - no risk"]
    }},
    "operation_type": "visualization",
    "sub_operation": "bar_chart",
    "source_columns": ["Product", "Sales"],
    "needs_new_column": false,
    "suggested_column_names": null,
    "operation_description": "Create bar chart showing Sales by Product",
    "formula_logic": "bar chart of Sales by Product",
    "is_ambiguous": false,
    "clarification_needed": null,
    "risk_level": "low",
    "execution_confidence": "high",
    "alternative_interpretations": []
}}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

NOW PARSE THE USER QUERY AND RETURN ONLY THE JSON:
"""
        
        return prompt
    
    def _extract_json_from_response(self, response) -> Dict:
        """
        Robust JSON extraction from AI response with multiple fallback strategies
        """
        try:
            content = response.choices[0].message.content.strip()
            
            # Strategy 1: Direct JSON parsing
            try:
                return json.loads(content)
            except json.JSONDecodeError:
                pass
            
            # Strategy 2: Remove markdown code blocks
            if '```' in content:
                # Extract content between code blocks
                parts = content.split('```')
                for part in parts:
                    part = part.strip()
                    if part.startswith('json'):
                        part = part[4:].strip()
                    if part and (part.startswith('{') or part.startswith('[')):
                        try:
                            return json.loads(part)
                        except json.JSONDecodeError:
                            continue
            
            # Strategy 3: Regex extraction of JSON object
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', content, re.DOTALL)
            if json_match:
                try:
                    return json.loads(json_match.group())
                except json.JSONDecodeError:
                    pass
            
            # Strategy 4: Find largest JSON-like structure
            brace_count = 0
            start_idx = -1
            for i, char in enumerate(content):
                if char == '{':
                    if brace_count == 0:
                        start_idx = i
                    brace_count += 1
                elif char == '}':
                    brace_count -= 1
                    if brace_count == 0 and start_idx != -1:
                        try:
                            return json.loads(content[start_idx:i+1])
                        except json.JSONDecodeError:
                            start_idx = -1
            
            # If all strategies fail, return error with context
            return {
                "error": "Failed to parse AI response into JSON",
                "raw_response": content[:500],  # Truncate for readability
                "parse_error": "Multiple parsing strategies failed",
                "suggestions": [
                    "AI may not have followed JSON format",
                    "Response may be ambiguous or incomplete",
                    "Try rephrasing your query"
                ]
            }
            
        except Exception as e:
            return {
                "error": "Unexpected error during response parsing",
                "exception": str(e),
                "type": type(e).__name__
            }
    
    def _validate_decision_plan(self, decision: AgentDecision, excel_structure: Dict) -> Dict:
        """
        Validate decision plan before execution with AI-assisted validation
        
        Returns:
            Validation report with approval status and recommendations
        """
        validation_prompt = f"""You are an Enterprise Risk Assessor validating an Excel operation plan.

PROPOSED OPERATION PLAN:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{json.dumps(decision.to_dict(), indent=2)}

EXCEL FILE CONTEXT:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Total Rows: {excel_structure.get('total_rows')}
Total Columns: {excel_structure.get('total_columns')}
Available Columns: {', '.join(excel_structure.get('columns', []))}
Numeric Columns: {', '.join(excel_structure.get('numeric_columns', []))}
Text Columns: {', '.join(excel_structure.get('text_columns', []))}

YOUR TASK:
Validate this decision plan and return ONLY a JSON assessment:

{{
    "is_valid": true,
    "validation_status": "approved | needs_review | rejected",
    "validation_checks": {{
        "columns_exist": true,
        "operation_feasible": true,
        "risk_acceptable": true,
        "data_types_compatible": true
    }},
    "warnings": [],
    "blockers": [],
    "recommendations": [],
    "requires_user_confirmation": false,
    "suggested_modifications": {{}}
}}

VALIDATION CRITERIA:
1. All source_columns must exist in available columns
2. Data types must be compatible with operation (e.g., can't sum text columns)
3. Risk level must be accurately assessed
4. For high-risk operations, flag for user confirmation
5. Check for potential data loss or corruption
6. Verify operation type matches sub_operation

Return ONLY the JSON validation response (no markdown, no explanations).
"""
        
        messages = [
            {"role": "system", "content": self.system_prompt},
            {"role": "user", "content": validation_prompt}
        ]
        
        response = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.2,
            max_tokens=800
        )
        
        return self._extract_json_from_response(response)
    
    def _generate_result_explanation(self, result: Dict, decision: AgentDecision) -> Dict:
        """
        Generate business-friendly explanation of operation results
        
        Returns:
            Professional summary for end users
        """
        explanation_prompt = f"""You are an Enterprise Communication Specialist explaining Excel operation results.

OPERATION EXECUTED:
{decision.change_description}

TECHNICAL RESULT:
{json.dumps(result, indent=2)}

YOUR TASK:
Create a clear, professional explanation for business users.

Return ONLY JSON:
{{
    "executive_summary": "One-sentence summary of what was accomplished",
    "what_changed": "Detailed explanation of changes made to the Excel file",
    "business_impact": "How this affects data analysis and decision-making",
    "next_steps": ["Recommended action 1", "Recommended action 2"],
    "technical_details": {{
        "rows_affected": 0,
        "columns_added": [],
        "file_location": "path/to/file.xlsx"
    }},
    "success_metrics": {{
        "operation_status": "completed",
        "data_quality": "validated",
        "audit_trail": "logged"
    }}
}}

Style: Professional, clear, non-technical language suitable for business stakeholders.
Return ONLY the JSON (no markdown).
"""
        
        messages = [
            {"role": "system", "content": self.system_prompt},
            {"role": "user", "content": explanation_prompt}
        ]
        
        response = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.3,
            max_tokens=1000
        )
        
        return self._extract_json_from_response(response)
    
    def _handle_error_recovery(self, error: str, query: str, context: Dict) -> Dict:
        """
        AI-assisted error recovery and suggestion generation
        
        Returns:
            Recovery strategies and alternative approaches
        """
        recovery_prompt = f"""You are an Enterprise Problem Solver for Excel operations.

ERROR ENCOUNTERED:
{error}

ORIGINAL USER QUERY:
"{query}"

CONTEXT:
{json.dumps(context, indent=2)}

YOUR TASK:
Analyze the error and suggest recovery strategies.

Return ONLY JSON:
{{
    "error_category": "data_issue | syntax_error | missing_column | type_mismatch | permission_error | other",
    "root_cause": "Brief explanation of what went wrong",
    "user_friendly_message": "Non-technical explanation for the user",
    "suggested_fixes": [
        {{
            "fix_description": "What to try",
            "modified_query": "Corrected query suggestion",
            "success_likelihood": "high | medium | low"
        }}
    ],
    "alternative_approaches": [
        "Alternative approach 1",
        "Alternative approach 2"
    ],
    "prevention_tips": ["How to avoid this error in the future"]
}}

Be helpful, solution-oriented, and empathetic. Return ONLY the JSON.
"""
        
        messages = [
            {"role": "system", "content": self.system_prompt},
            {"role": "user", "content": recovery_prompt}
        ]
        
        response = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.4,
            max_tokens=1000
        )
        
        return self._extract_json_from_response(response)
    
    def create_decision_plan(self, parsed_query: Dict, excel_structure: Dict) -> AgentDecision:
        """
        Creates an enterprise-grade structured decision plan based on parsed query
        
        Returns:
            AgentDecision object with complete execution plan
        """
        # Extract information with null safety
        operation_type = parsed_query.get("operation_type", "calculation") or "calculation"
        sub_operation = parsed_query.get("sub_operation", "sum") or "sum"
        source_columns = parsed_query.get("source_columns", []) or []
        
        # Ensure source_columns contains only strings (filter out None values)
        source_columns = [str(col) for col in source_columns if col is not None]
        
        # Handle suggested column names with null safety
        suggested_names = parsed_query.get("suggested_column_names")
        if suggested_names is None:
            target_columns = ["Result"]
        elif isinstance(suggested_names, list):
            # Filter out None values from the list
            target_columns = [str(name) for name in suggested_names if name is not None]
            if not target_columns:
                target_columns = ["Result"]
        else:
            target_columns = [str(suggested_names)]
        
        operation_desc = parsed_query.get("operation_description", "") or ""
        formula_logic = parsed_query.get("formula_logic", "") or ""
        sheet_name = excel_structure.get("analyzed_sheet", "Sheet1") or "Sheet1"
        
        # Check if this is a read-only operation (display/view)
        is_read_only = False
        if operation_type == "reporting" and sub_operation in ["data_display", "view", "show", "display"]:
            is_read_only = True
        
        # For visualization and some operations, no target column needed
        if operation_type in ["visualization", "reporting"]:
            target_columns = None
        
        # Determine risk level based on operation type and parsed query
        risk_level = parsed_query.get("risk_level", "low") or "low"
        if operation_type == "transformation" and sub_operation in ["remove_column", "deduplicate"]:
            risk_level = "high"
        elif operation_type in ["formatting", "validation"]:
            risk_level = "medium"
        
        # Requires confirmation for high-risk operations
        requires_confirmation = (risk_level == "high")
        
        # Helper function to safely join columns
        def safe_join(columns, separator=", "):
            if not columns:
                return "N/A"
            valid_cols = [str(c) for c in columns if c is not None]
            return separator.join(valid_cols) if valid_cols else "N/A"
        
        # Create change description in business language
        source_cols_str = safe_join(source_columns)
        target_cols_str = safe_join(target_columns) if target_columns else "N/A"
        
        if operation_type == "calculation":
            change_desc = f"Calculate {sub_operation} of {source_cols_str} and create new column(s): {target_cols_str}"
        elif operation_type == "visualization":
            change_desc = f"Create {sub_operation} chart using {source_cols_str}"
        elif operation_type == "transformation":
            change_desc = f"Transform data: {sub_operation} on {source_cols_str}"
        elif operation_type == "aggregation":
            change_desc = f"Aggregate data using {sub_operation} on {source_cols_str}"
        elif operation_type == "lookup":
            change_desc = f"Perform {sub_operation} to find matching data in {source_cols_str}"
        elif operation_type == "formatting":
            change_desc = f"Apply {sub_operation} formatting to {source_cols_str}"
        elif operation_type == "validation":
            change_desc = f"Add {sub_operation} validation rules to {source_cols_str}"
        elif operation_type == "reporting" and is_read_only:
            change_desc = f"Display data from {source_cols_str} (read-only, no file modification)"
        else:
            change_desc = operation_desc if operation_desc else f"{operation_type}: {sub_operation}"
        
        # Create execution scope-specific assumptions
        execution_scope = "all_rows"
        total_rows = excel_structure.get('total_rows', 0)
        
        # Build assumptions list
        assumptions = []
        if target_columns:
            assumptions.append(f"Will create new column(s): {target_cols_str}")
        if is_read_only:
            assumptions.append("Read-only operation - no file will be modified")
        else:
            assumptions.append("Original data will remain unchanged (new file created)")
        assumptions.append(f"Operation will apply to {execution_scope} ({total_rows} rows)")
        if operation_type == "visualization":
            assumptions.append("Chart will be embedded in the Excel file below the data")
        
        # Create decision with all fields including new ones
        decision = AgentDecision(
            operation_type=operation_type,
            sub_operation=sub_operation,
            source_columns=source_columns,
            target_columns=target_columns,
            sheet_name=sheet_name,
            execution_scope=execution_scope,
            risk_level=risk_level,
            assumptions=assumptions,
            requires_confirmation=requires_confirmation,
            change_description=change_desc,
            formula_logic=formula_logic,
            is_read_only=is_read_only
        )
        
        return decision
    
    def _save_with_all_sheets(self, file_path: str, modified_df: pd.DataFrame, target_sheet: str, output_path: str):
        """
        Save the modified dataframe while preserving all other sheets from the original file.
        
        Args:
            file_path: Original Excel file path
            modified_df: Modified dataframe to save
            target_sheet: Name of the sheet that was modified
            output_path: Output file path
        """
        try:
            # Load all sheets from original file
            with pd.ExcelFile(file_path) as xls:
                all_sheets = {}
                for sheet in xls.sheet_names:
                    if sheet == target_sheet:
                        all_sheets[sheet] = modified_df
                    else:
                        all_sheets[sheet] = pd.read_excel(xls, sheet_name=sheet)
            
            # Save all sheets to output using ExcelWriter
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, sheet_df in all_sheets.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
        except Exception as e:
            # Fallback: save just the modified sheet
            modified_df.to_excel(output_path, sheet_name=target_sheet, index=False)
    
    def execute_excel_operation(self, file_path: str, decision: AgentDecision, output_path: str) -> Dict:
        """
        Executes the planned Excel operation (Enterprise-grade)
        Preserves all sheets when saving.
        
        Returns:
            Dictionary with execution status and details
        """
        try:
            # Check for read-only operations
            if decision.is_read_only:
                df = pd.read_excel(file_path, sheet_name=decision.sheet_name)
                return {
                    "status": "success",
                    "message": decision.change_description,
                    "rows_affected": len(df),
                    "output_file": None,
                    "changes": decision.assumptions,
                    "operation": f"{decision.operation_type} - {decision.sub_operation}",
                    "is_read_only": True,
                    "data_preview": df.head(10).to_dict('records')
                }
            
            # Load Excel file
            df = pd.read_excel(file_path, sheet_name=decision.sheet_name)
            
            # Execute based on operation type
            if decision.operation_type == "calculation":
                df = self._execute_calculation(df, decision)
                self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
                
                return {
                    "status": "success",
                    "message": decision.change_description,
                    "rows_affected": len(df),
                    "output_file": output_path,
                    "changes": decision.assumptions,
                    "operation": f"{decision.operation_type} - {decision.sub_operation}"
                }
            
            elif decision.operation_type == "transformation":
                df = self._execute_transformation(df, decision)
                self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
                
                return {
                    "status": "success",
                    "message": decision.change_description,
                    "rows_affected": len(df),
                    "output_file": output_path,
                    "changes": decision.assumptions,
                    "operation": f"{decision.operation_type} - {decision.sub_operation}"
                }
            
            elif decision.operation_type == "aggregation":
                df = self._execute_aggregation(df, decision)
                self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
                
                return {
                    "status": "success",
                    "message": decision.change_description,
                    "rows_affected": len(df),
                    "output_file": output_path,
                    "changes": decision.assumptions,
                    "operation": f"{decision.operation_type} - {decision.sub_operation}"
                }
            
            elif decision.operation_type == "formatting":
                return self._execute_formatting(file_path, df, decision, output_path)
            
            elif decision.operation_type == "visualization":
                return self._execute_visualization(file_path, df, decision, output_path)
            
            elif decision.operation_type == "validation":
                return self._execute_validation(file_path, df, decision, output_path)
            
            elif decision.operation_type == "lookup":
                df = self._execute_lookup(df, decision, file_path=file_path)
                self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
                
                return {
                    "status": "success",
                    "message": decision.change_description,
                    "rows_affected": len(df),
                    "output_file": output_path,
                    "changes": decision.assumptions,
                    "operation": f"{decision.operation_type} - {decision.sub_operation}"
                }
            
            else:
                # Default: save as is (preserving all sheets)
                self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
                
                return {
                    "status": "success",
                    "message": f"Successfully processed {decision.operation_type}",
                    "rows_affected": len(df),
                    "output_file": output_path,
                    "changes": decision.assumptions,
                    "operation": decision.operation_type
                }
            
        except Exception as e:
            return {
                "status": "failed",
                "error": str(e),
                "operation": f"{decision.operation_type} - {decision.sub_operation}"
            }
    
    def _execute_calculation(self, df: pd.DataFrame, decision: AgentDecision) -> pd.DataFrame:
        """Execute calculation operations (Enterprise-grade)"""
        source_cols = decision.source_columns
        target_cols = decision.target_columns if decision.target_columns else ["Result"]
        target_col = target_cols[0]  # Primary target column
        sub_op = decision.sub_operation.lower()
        
        # CALCULATION SUB-OPERATIONS
        if sub_op == "sum":
            df[target_col] = df[source_cols].sum(axis=1)
        
        elif sub_op == "average" or sub_op == "mean":
            df[target_col] = df[source_cols].mean(axis=1)
        
        elif sub_op == "min" or sub_op == "minimum":
            df[target_col] = df[source_cols].min(axis=1)
        
        elif sub_op == "max" or sub_op == "maximum":
            df[target_col] = df[source_cols].max(axis=1)
        
        elif sub_op == "count":
            df[target_col] = df[source_cols].count(axis=1)
        
        elif sub_op == "median":
            df[target_col] = df[source_cols].median(axis=1)
        
        elif sub_op == "multiply":
            result = df[source_cols[0]]
            for col in source_cols[1:]:
                result = result * df[col]
            df[target_col] = result
        
        elif sub_op == "divide":
            if len(source_cols) >= 2:
                df[target_col] = df[source_cols[0]] / df[source_cols[1]]
        
        elif sub_op == "percentage":
            if len(source_cols) >= 2:
                df[target_col] = (df[source_cols[0]] / df[source_cols[1]]) * 100
            else:
                # Percentage of total
                df[target_col] = (df[source_cols[0]] / df[source_cols[0]].sum()) * 100
        
        else:
            # Default: sum
            df[target_col] = df[source_cols].sum(axis=1)
        
        return df
    
    def _execute_transformation(self, df: pd.DataFrame, decision: AgentDecision) -> pd.DataFrame:
        """Execute data transformation operations"""
        sub_op = decision.sub_operation.lower()
        source_cols = decision.source_columns
        target_cols = decision.target_columns
        
        if sub_op == "add_column":
            # Add new empty column(s)
            if target_cols:
                for col in target_cols:
                    df[col] = None
        
        elif sub_op == "rename_column":
            # Rename column (source[0] -> target[0])
            if source_cols and target_cols:
                df.rename(columns={source_cols[0]: target_cols[0]}, inplace=True)
        
        elif sub_op == "remove_column":
            # Remove columns (high risk - requires confirmation)
            df.drop(columns=source_cols, inplace=True, errors='ignore')
        
        elif sub_op == "sort":
            # Sort by columns
            df.sort_values(by=source_cols, inplace=True)
        
        elif sub_op == "deduplicate":
            # Remove duplicates based on columns
            df.drop_duplicates(subset=source_cols, inplace=True)
        
        elif sub_op == "split_column":
            # Split column (basic implementation - can be enhanced)
            if source_cols and target_cols and len(target_cols) >= 2:
                # Split first source column by space/comma
                split_data = df[source_cols[0]].str.split(expand=True)
                for idx, target_col in enumerate(target_cols[:split_data.shape[1]]):
                    df[target_col] = split_data[idx]
        
        elif sub_op == "merge_columns":
            # Merge multiple columns into one
            if source_cols and target_cols:
                df[target_cols[0]] = df[source_cols].apply(lambda row: ' '.join(row.values.astype(str)), axis=1)
        
        elif sub_op == "fill_missing" or sub_op == "replace_nan" or sub_op == "fillna":
            # Fill missing/NaN values with a specified value
            # Get the fill value from formula_logic or use a default
            fill_value = decision.formula_logic if decision.formula_logic else "N/A"
            
            # If specific columns are provided, fill only those
            if source_cols:
                for col in source_cols:
                    if col in df.columns:
                        df[col] = df[col].fillna(fill_value)
            else:
                # Fill all columns
                df = df.fillna(fill_value)
        
        elif sub_op == "replace_values":
            # Replace specific values in columns
            # formula_logic should contain: "old_value->new_value"
            if decision.formula_logic and "->" in decision.formula_logic:
                old_val, new_val = decision.formula_logic.split("->", 1)
                old_val = old_val.strip()
                new_val = new_val.strip()
                
                if source_cols:
                    for col in source_cols:
                        if col in df.columns:
                            df[col] = df[col].replace(old_val, new_val)
                else:
                    df = df.replace(old_val, new_val)
        
        elif sub_op == "filter":
            # Filter rows based on condition
            # formula_logic should contain condition like: "column_name > 100" or "column_name == 'value'"
            if decision.formula_logic:
                condition = decision.formula_logic.strip()
                try:
                    # Parse common filter patterns
                    # Pattern: column_name operator value
                    operators = ['>=', '<=', '!=', '==', '>', '<', 'contains', 'startswith', 'endswith']
                    
                    for op in operators:
                        if op in condition:
                            parts = condition.split(op, 1)
                            if len(parts) == 2:
                                col_name = parts[0].strip()
                                value = parts[1].strip().strip('"').strip("'")
                                
                                if col_name in df.columns:
                                    if op == 'contains':
                                        df = df[df[col_name].astype(str).str.contains(value, case=False, na=False)]
                                    elif op == 'startswith':
                                        df = df[df[col_name].astype(str).str.startswith(value, na=False)]
                                    elif op == 'endswith':
                                        df = df[df[col_name].astype(str).str.endswith(value, na=False)]
                                    elif op == '>=':
                                        df = df[df[col_name] >= float(value)]
                                    elif op == '<=':
                                        df = df[df[col_name] <= float(value)]
                                    elif op == '>':
                                        df = df[df[col_name] > float(value)]
                                    elif op == '<':
                                        df = df[df[col_name] < float(value)]
                                    elif op == '==':
                                        # Try numeric first, then string
                                        try:
                                            df = df[df[col_name] == float(value)]
                                        except:
                                            df = df[df[col_name] == value]
                                    elif op == '!=':
                                        try:
                                            df = df[df[col_name] != float(value)]
                                        except:
                                            df = df[df[col_name] != value]
                                break
                except Exception as e:
                    print(f"   ⚠️  Filter parsing error: {e}")
        
        elif sub_op == "sort_desc" or sub_op == "sort_descending":
            # Sort by columns in descending order
            if source_cols:
                df.sort_values(by=source_cols, ascending=False, inplace=True)
        
        return df
    
    def _execute_lookup(self, df: pd.DataFrame, decision: AgentDecision, file_path: str = None) -> pd.DataFrame:
        """
        Execute lookup/reference operations including cross-sheet lookups.
        
        Supports:
        - vlookup: Look up value in another column/sheet
        - xlookup: Modern lookup with more flexibility
        - index_match: INDEX-MATCH combination
        """
        sub_op = decision.sub_operation.lower()
        source_cols = decision.source_columns
        target_cols = decision.target_columns
        formula_logic = decision.formula_logic or ""
        
        if sub_op in ["vlookup", "xlookup", "lookup"]:
            # formula_logic format: "lookup_col:return_col:sheet_name" or "lookup_col:return_col"
            if formula_logic and ":" in formula_logic:
                parts = formula_logic.split(":")
                if len(parts) >= 2:
                    lookup_col = parts[0].strip()
                    return_col = parts[1].strip()
                    ref_sheet = parts[2].strip() if len(parts) > 2 else decision.sheet_name
                    
                    # Load reference sheet
                    if file_path:
                        try:
                            ref_df = pd.read_excel(file_path, sheet_name=ref_sheet)
                            
                            # Perform lookup
                            if lookup_col in df.columns and lookup_col in ref_df.columns and return_col in ref_df.columns:
                                # Create lookup dictionary
                                lookup_dict = dict(zip(ref_df[lookup_col], ref_df[return_col]))
                                
                                # Create new column with looked up values
                                new_col_name = target_cols[0] if target_cols else f"{return_col}_lookup"
                                df[new_col_name] = df[lookup_col].map(lookup_dict)
                        except Exception as e:
                            print(f"   ⚠️  Lookup error: {e}")
        
        elif sub_op == "index_match":
            # Similar to vlookup but with more flexibility
            if source_cols and len(source_cols) >= 2 and target_cols:
                match_col = source_cols[0]
                return_col = source_cols[1]
                
                if match_col in df.columns and return_col in df.columns:
                    # Create mapping
                    lookup_dict = dict(zip(df[match_col], df[return_col]))
                    df[target_cols[0]] = df[match_col].map(lookup_dict)
        
        elif sub_op == "merge" or sub_op == "join_sheets":
            # Merge with another sheet
            # formula_logic format: "sheet_name:join_column:join_type"
            if formula_logic and ":" in formula_logic and file_path:
                parts = formula_logic.split(":")
                if len(parts) >= 2:
                    ref_sheet = parts[0].strip()
                    join_col = parts[1].strip()
                    join_type = parts[2].strip() if len(parts) > 2 else "left"
                    
                    try:
                        ref_df = pd.read_excel(file_path, sheet_name=ref_sheet)
                        if join_col in df.columns and join_col in ref_df.columns:
                            df = df.merge(ref_df, on=join_col, how=join_type)
                    except Exception as e:
                        print(f"   ⚠️  Merge error: {e}")
        
        return df
    
    def _execute_validation(self, file_path: str, df: pd.DataFrame, decision: AgentDecision, output_path: str) -> Dict:
        """
        Execute data validation operations using openpyxl.
        
        Supports:
        - dropdown: Create dropdown list from values
        - range_check: Validate numeric ranges
        - text_length: Validate text length
        - custom: Custom validation formula
        """
        from openpyxl.worksheet.datavalidation import DataValidation
        
        try:
            # Save dataframe first (preserving all sheets)
            self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
            
            # Load with openpyxl for validation rules
            wb = openpyxl.load_workbook(output_path)
            ws = wb[decision.sheet_name]
            
            sub_op = decision.sub_operation.lower()
            source_cols = decision.source_columns
            formula_logic = decision.formula_logic or ""
            
            # Get column letters for source columns
            col_positions = {col: idx + 1 for idx, col in enumerate(df.columns)}
            
            validation_applied = []
            
            for col in source_cols:
                if col in col_positions:
                    col_letter = openpyxl.utils.get_column_letter(col_positions[col])
                    cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
                    
                    if sub_op == "dropdown":
                        # Create dropdown from formula_logic values (comma-separated)
                        if formula_logic:
                            options = formula_logic.replace(";", ",")
                            dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
                            dv.error = "Please select from the dropdown list"
                            dv.errorTitle = "Invalid Input"
                            dv.prompt = "Select a value from the list"
                            dv.promptTitle = "Dropdown Selection"
                            ws.add_data_validation(dv)
                            dv.add(cell_range)
                            validation_applied.append(f"Dropdown on {col}: {options}")
                    
                    elif sub_op == "range_check" or sub_op == "numeric_range":
                        # Validate numeric range: formula_logic = "min:max"
                        if formula_logic and ":" in formula_logic:
                            parts = formula_logic.split(":")
                            min_val = parts[0].strip()
                            max_val = parts[1].strip() if len(parts) > 1 else None
                            
                            if min_val and max_val:
                                dv = DataValidation(type="decimal", operator="between",
                                                   formula1=min_val, formula2=max_val)
                            elif min_val:
                                dv = DataValidation(type="decimal", operator="greaterThanOrEqual",
                                                   formula1=min_val)
                            dv.error = f"Value must be between {min_val} and {max_val}"
                            dv.errorTitle = "Invalid Value"
                            ws.add_data_validation(dv)
                            dv.add(cell_range)
                            validation_applied.append(f"Range check on {col}: {min_val} to {max_val}")
                    
                    elif sub_op == "text_length":
                        # Validate text length: formula_logic = "max_length"
                        if formula_logic:
                            max_len = formula_logic.strip()
                            dv = DataValidation(type="textLength", operator="lessThanOrEqual",
                                               formula1=max_len)
                            dv.error = f"Text must be {max_len} characters or less"
                            dv.errorTitle = "Text Too Long"
                            ws.add_data_validation(dv)
                            dv.add(cell_range)
                            validation_applied.append(f"Text length on {col}: max {max_len}")
                    
                    elif sub_op == "no_duplicates" or sub_op == "unique":
                        # Mark duplicates (validation can't prevent, but we can highlight)
                        dv = DataValidation(type="custom", formula1=f'=COUNTIF({col_letter}:{col_letter},{col_letter}2)=1')
                        dv.error = "Duplicate value detected"
                        dv.errorTitle = "Duplicate Entry"
                        ws.add_data_validation(dv)
                        dv.add(cell_range)
                        validation_applied.append(f"Unique check on {col}")
            
            wb.save(output_path)
            
            return {
                "status": "success",
                "message": f"Validation rules applied: {', '.join(validation_applied) if validation_applied else decision.sub_operation}",
                "output_file": output_path,
                "rows_affected": len(df),
                "changes": decision.assumptions + validation_applied,
                "operation": f"{decision.operation_type} - {decision.sub_operation}"
            }
        except Exception as e:
            return {
                "status": "failed",
                "error": f"Validation failed: {str(e)}"
            }
    
    def _execute_aggregation(self, df: pd.DataFrame, decision: AgentDecision) -> pd.DataFrame:
        """
        Execute aggregation operations.
        
        Supports:
        - group_by: Group data and aggregate
        - pivot_table: Create pivot table
        - summary_stats: Generate summary statistics
        - subtotals: Add subtotal rows
        """
        sub_op = decision.sub_operation.lower()
        source_cols = decision.source_columns
        target_cols = decision.target_columns
        formula_logic = decision.formula_logic or "sum"
        
        if sub_op == "group_by" or sub_op == "groupby":
            # formula_logic format: "agg_column:agg_function" or just "agg_function"
            if source_cols:
                group_cols = source_cols[:-1] if len(source_cols) > 1 else source_cols
                agg_col = source_cols[-1] if len(source_cols) > 1 else None
                
                # Parse aggregation function from formula_logic
                agg_func = formula_logic.lower() if formula_logic else "sum"
                if agg_func not in ["sum", "mean", "count", "min", "max", "median", "std"]:
                    agg_func = "sum"
                
                if agg_col and agg_col in df.columns:
                    # Group by and aggregate
                    grouped = df.groupby(group_cols, as_index=False).agg({agg_col: agg_func})
                    df = grouped
                else:
                    # Just group and count
                    grouped = df.groupby(group_cols, as_index=False).size()
                    grouped.columns = list(group_cols) + ['Count']
                    df = grouped
        
        elif sub_op == "pivot_table" or sub_op == "pivot":
            # formula_logic format: "values_col:agg_func"
            if len(source_cols) >= 2:
                index_col = source_cols[0]
                columns_col = source_cols[1] if len(source_cols) > 1 else None
                values_col = source_cols[2] if len(source_cols) > 2 else None
                
                agg_func = formula_logic.lower() if formula_logic else "sum"
                
                if values_col and values_col in df.columns:
                    pivot = pd.pivot_table(df, values=values_col, index=index_col,
                                          columns=columns_col, aggfunc=agg_func, fill_value=0)
                    df = pivot.reset_index()
                else:
                    # Simple pivot with count
                    pivot = pd.pivot_table(df, index=index_col, columns=columns_col,
                                          aggfunc='size', fill_value=0)
                    df = pivot.reset_index()
        
        elif sub_op == "summary_stats" or sub_op == "describe":
            # Generate summary statistics for numeric columns
            if source_cols:
                numeric_cols = [c for c in source_cols if c in df.select_dtypes(include=['number']).columns]
                if numeric_cols:
                    summary = df[numeric_cols].describe().T
                    summary = summary.reset_index()
                    summary.columns = ['Column', 'Count', 'Mean', 'Std', 'Min', '25%', '50%', '75%', 'Max']
                    df = summary
            else:
                df = df.describe().T.reset_index()
        
        elif sub_op == "subtotals":
            # Add subtotal rows for grouped data
            if source_cols:
                group_col = source_cols[0]
                numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
                
                if group_col in df.columns and numeric_cols:
                    # Calculate subtotals
                    subtotals = df.groupby(group_col)[numeric_cols].sum()
                    subtotals['_is_subtotal'] = True
                    
                    # Interleave with original data
                    result_dfs = []
                    for group_val in df[group_col].unique():
                        group_data = df[df[group_col] == group_val].copy()
                        group_data['_is_subtotal'] = False
                        result_dfs.append(group_data)
                        
                        # Add subtotal row
                        subtotal_row = subtotals.loc[group_val].to_frame().T
                        subtotal_row[group_col] = f"{group_val} - SUBTOTAL"
                        subtotal_row['_is_subtotal'] = True
                        result_dfs.append(subtotal_row)
                    
                    df = pd.concat(result_dfs, ignore_index=True)
        
        elif sub_op == "value_counts" or sub_op == "frequency":
            # Count frequency of values in a column
            if source_cols and source_cols[0] in df.columns:
                counts = df[source_cols[0]].value_counts().reset_index()
                counts.columns = [source_cols[0], 'Count']
                df = counts
        
        return df
    
    def _execute_formatting(self, file_path: str, df: pd.DataFrame, decision: AgentDecision, output_path: str) -> Dict:
        """
        Execute formatting operations using openpyxl.
        
        Supports:
        - conditional_color: Color cells based on conditions
        - highlight: Highlight specific values
        - data_bars: Add data bar visualization
        - color_scale: Apply color scale (heatmap)
        """
        from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
        from openpyxl.styles import PatternFill, Font, Border, Side
        
        try:
            # Save dataframe first (preserving all sheets)
            self._save_with_all_sheets(file_path, df, decision.sheet_name, output_path)
            
            # Load workbook
            wb = openpyxl.load_workbook(output_path)
            ws = wb[decision.sheet_name]
            
            sub_op = decision.sub_operation.lower()
            source_cols = decision.source_columns
            formula_logic = decision.formula_logic or ""
            
            # Get column positions
            col_positions = {col: idx + 1 for idx, col in enumerate(df.columns)}
            
            formatting_applied = []
            
            for col in source_cols:
                if col in col_positions:
                    col_letter = openpyxl.utils.get_column_letter(col_positions[col])
                    cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"
                    
                    if sub_op == "highlight_negative" or sub_op == "highlight_negatives":
                        # Highlight negative values in red
                        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        rule = FormulaRule(formula=[f'{col_letter}2<0'], fill=red_fill)
                        ws.conditional_formatting.add(cell_range, rule)
                        formatting_applied.append(f"Highlighted negatives in {col}")
                    
                    elif sub_op == "highlight_positive" or sub_op == "highlight_positives":
                        # Highlight positive values in green
                        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                        rule = FormulaRule(formula=[f'{col_letter}2>0'], fill=green_fill)
                        ws.conditional_formatting.add(cell_range, rule)
                        formatting_applied.append(f"Highlighted positives in {col}")
                    
                    elif sub_op == "highlight" or sub_op == "conditional_color":
                        # Highlight based on condition in formula_logic: ">100" or "==value"
                        if formula_logic:
                            # Parse condition
                            condition = formula_logic.strip()
                            fill_color = "FFFF00"  # Default yellow
                            
                            # Check for color specification: ">100:FF0000"
                            if ":" in condition and len(condition.split(":")[-1]) == 6:
                                parts = condition.rsplit(":", 1)
                                condition = parts[0]
                                fill_color = parts[1]
                            
                            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                            formula = f'{col_letter}2{condition}'
                            rule = FormulaRule(formula=[formula], fill=fill)
                            ws.conditional_formatting.add(cell_range, rule)
                            formatting_applied.append(f"Conditional formatting on {col}: {condition}")
                    
                    elif sub_op == "data_bars":
                        # Add data bars
                        rule = DataBarRule(start_type='min', end_type='max',
                                          color="638EC6", showValue=True, minLength=None, maxLength=None)
                        ws.conditional_formatting.add(cell_range, rule)
                        formatting_applied.append(f"Data bars on {col}")
                    
                    elif sub_op == "color_scale" or sub_op == "heatmap":
                        # Apply color scale (red-yellow-green)
                        rule = ColorScaleRule(start_type='min', start_color='F8696B',
                                             mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                             end_type='max', end_color='63BE7B')
                        ws.conditional_formatting.add(cell_range, rule)
                        formatting_applied.append(f"Color scale on {col}")
                    
                    elif sub_op == "bold_header":
                        # Make header bold
                        ws[f'{col_letter}1'].font = Font(bold=True)
                        formatting_applied.append(f"Bold header for {col}")
            
            # Apply header formatting to all columns if no specific columns
            if sub_op == "format_headers" or sub_op == "style_headers":
                for col_idx in range(1, len(df.columns) + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws[f'{col_letter}1'].font = Font(bold=True)
                    ws[f'{col_letter}1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    ws[f'{col_letter}1'].font = Font(bold=True, color="FFFFFF")
                formatting_applied.append("Formatted all headers")
            
            wb.save(output_path)
            
            return {
                "status": "success",
                "message": f"Formatting applied: {', '.join(formatting_applied) if formatting_applied else 'Formatting complete'}",
                "output_file": output_path,
                "rows_affected": len(df),
                "changes": decision.assumptions + formatting_applied,
                "operation": f"{decision.operation_type} - {decision.sub_operation}"
            }
            
        except Exception as e:
            return {
                "status": "failed",
                "error": f"Formatting failed: {str(e)}"
            }
    
    def _execute_visualization(self, file_path: str, df: pd.DataFrame, decision: AgentDecision, output_path: str) -> Dict:
        """Execute visualization operations - create charts in Excel"""
        try:
            # First save the data to Excel
            df.to_excel(output_path, sheet_name=decision.sheet_name, index=False)
            
            # Load the workbook to add chart
            wb = openpyxl.load_workbook(output_path)
            ws = wb[decision.sheet_name]
            
            # Determine chart type from formula logic
            formula_lower = decision.  formula_logic.lower()
            source_cols = decision.source_columns
            
            # Identify category column (usually text/product name) and data columns (numeric)
            category_col = None
            data_cols = []
            
            for col in source_cols:
                if col in df.columns:
                    if df[col].dtype == 'object':  # Text column
                        category_col = col
                    else:  # Numeric column
                        data_cols.append(col)
            
            # If no category column found, use first column
            if not category_col and len(source_cols) > 0:
                category_col = source_cols[0]
                data_cols = source_cols[1:] if len(source_cols) > 1 else source_cols
            
            # Get column positions in Excel (1-indexed)
            col_positions = {col: idx + 1 for idx, col in enumerate(df.columns)}
            
            # Determine chart type
            if "pie" in formula_lower:
                chart = PieChart()
                chart.title = "Pie Chart"
            elif "line" in formula_lower:
                chart = LineChart()
                chart.title = "Line Chart"
            else:
                # Default to bar chart
                chart = BarChart()
                chart.title = "Bar Chart"
            
            # Set chart dimensions
            chart.height = 15
            chart.width = 25
            
            # Set data range (excluding header)
            min_row = 2
            max_row = len(df) + 1
            
            # Add data series
            for data_col in data_cols:
                if data_col in col_positions:
                    col_letter = openpyxl.utils.get_column_letter(col_positions[data_col])
                    
                    # Data values
                    data = Reference(ws, min_col=col_positions[data_col], min_row=1, 
                                   max_row=max_row)
                    chart.add_data(data, titles_from_data=True)
            
            # Set categories (x-axis labels)
            if category_col and category_col in col_positions:
                cat_col_letter = openpyxl.utils.get_column_letter(col_positions[category_col])
                cats = Reference(ws, min_col=col_positions[category_col], min_row=2, 
                               max_row=max_row)
                chart.set_categories(cats)
            
            # Set chart labels - only for charts that support axes (not PieChart)
            if hasattr(chart, 'x_axis') and hasattr(chart, 'y_axis'):
                chart.x_axis.title = category_col if category_col else "Categories"
                chart.y_axis.title = "Values"
            
            # Style
            chart.style = 10
            
            # Add chart to worksheet at a position below the data
            chart_position = f"A{max_row + 3}"
            ws.add_chart(chart, chart_position)
            
            # Save workbook with chart
            wb.save(output_path)
            
            return {
                "status": "success",
                "message": f"Successfully created bar chart for {', '.join(data_cols)}",
                "rows_affected": len(df),
                "output_file": output_path,
                "changes": [
                    f"Created {'pie' if 'pie' in formula_lower else 'line' if 'line' in formula_lower else 'bar'} chart",
                    f"Chart shows: {', '.join(data_cols)}",
                    f"Categories: {category_col}" if category_col else "No categories",
                    f"Chart added at position {chart_position}"
                ]
            }
            
        except Exception as e:
            return {
                "status": "failed",
                "error": f"Chart creation failed: {str(e)}"
            }
    
    def process_query(self, file_path: str, query: str, output_path: Optional[str] = None, sheet_name: Optional[str] = None, use_code_generation: bool = True) -> Dict:
        """
        Main entry point: Process user query on Excel file
        
        This is the orchestration method that follows the complete workflow
        
        Args:
            file_path: Path to the Excel file
            query: User's natural language query
            output_path: Optional path for output file
            sheet_name: Optional specific sheet to work with
            use_code_generation: If True, use LLM code generation (recommended)
        """
        print("\n" + "="*60)
        print("🤖 EXCEL AGENT - PROCESSING REQUEST")
        print("="*60)
        
        # Use code generation approach for intelligent execution
        if use_code_generation:
            return self._process_with_code_generation(file_path, query, output_path, sheet_name)
        
        # Legacy approach (kept for backward compatibility)
        return self._process_with_pattern_matching(file_path, query, output_path, sheet_name)
    
    def _process_with_code_generation(self, file_path: str, query: str, output_path: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict:
        """
        Process query using LLM code generation - the intelligent approach.
        LLM generates pandas code, we execute it safely.
        """
        # STEP 1: Analyze Excel Structure
        print("\n📊 STEP 1: Analyzing Excel Structure...")
        excel_structure = self.analyze_excel_structure(file_path, sheet_name=sheet_name)
        
        if "error" in excel_structure:
            return {"status": "failed", "error": excel_structure["error"]}
        
        target_sheet = excel_structure.get("analyzed_sheet", "Sheet1")
        columns_display = ', '.join([str(c) for c in excel_structure.get('columns', [])])
        print(f"   ✓ Found {excel_structure['total_columns']} columns: {columns_display}")
        print(f"   ✓ Total rows: {excel_structure['total_rows']}")
        print(f"   ✓ Sheet: {target_sheet}")
        
        # STEP 2: Generate pandas code using LLM
        print(f"\n🧠 STEP 2: AI Code Generation...")
        print(f"   Query: '{query}'")
        
        code_response = self._generate_pandas_code(query, excel_structure)
        
        if not code_response or "pandas_code" not in code_response:
            print("   ❌ Failed to generate code, falling back to pattern matching...")
            return self._process_with_pattern_matching(file_path, query, output_path, sheet_name)
        
        pandas_code = code_response.get("pandas_code", "")
        operation_desc = code_response.get("operation_description", "Execute operation")
        is_read_only = code_response.get("is_read_only", False)
        risk_level = code_response.get("risk_level", "medium")
        columns_affected = code_response.get("columns_affected", [])
        
        print(f"\n   💻 Generated Pandas Code:")
        for line in pandas_code.split('\n'):
            print(f"      {line}")
        print(f"\n   📋 Operation: {operation_desc}")
        print(f"   ⚠️  Risk Level: {risk_level.upper()}")
        print(f"   📊 Columns Affected: {', '.join(columns_affected) if columns_affected else 'All'}")
        
        # STEP 3: Load DataFrame
        print(f"\n⚙️  STEP 3: Executing Generated Code...")
        df = pd.read_excel(file_path, sheet_name=target_sheet)
        original_row_count = len(df)
        
        # STEP 4: Execute the generated code
        result_df, success, error = self._execute_generated_code(df, pandas_code, query)
        
        if not success:
            print(f"   ❌ Code execution failed: {error}")
            print("   🔄 Falling back to pattern matching approach...")
            return self._process_with_pattern_matching(file_path, query, output_path, sheet_name)
        
        new_row_count = len(result_df)
        print(f"   ✅ Code executed successfully!")
        print(f"   📊 Rows: {original_row_count} → {new_row_count}")
        
        # STEP 5: Handle results
        if is_read_only:
            print(f"\n📊 STEP 4: Displaying Results (Read-Only)...")
            return {
                "status": "success",
                "message": operation_desc,
                "rows_affected": new_row_count,
                "output_file": None,
                "is_read_only": True,
                "data_preview": self._make_json_safe(result_df.head(10).to_dict('records')),
                "generated_code": pandas_code
            }
        
        # Generate output filename
        print(f"\n💾 STEP 4: Saving Results...")
        if not output_path:
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            dir_name = os.path.dirname(file_path)
            counter = 1
            while True:
                output_path = os.path.join(dir_name, f"{base_name}_modified_{counter}.xlsx") if dir_name else f"{base_name}_modified_{counter}.xlsx"
                if not os.path.exists(output_path):
                    break
                counter += 1
        
        # Save with all sheets preserved
        self._save_with_all_sheets(file_path, result_df, target_sheet, output_path)
        
        print(f"   ✅ SUCCESS!")
        print(f"   ✓ Operation: {operation_desc}")
        print(f"   ✓ Rows: {original_row_count} → {new_row_count}")
        print(f"   ✓ Output saved to: {output_path}")
        
        # Log the change
        self.change_log.append({
            "timestamp": pd.Timestamp.now().isoformat(),
            "file": file_path,
            "query": query,
            "operation": operation_desc,
            "output": output_path,
            "generated_code": pandas_code
        })
        
        return {
            "status": "success",
            "message": operation_desc,
            "rows_before": original_row_count,
            "rows_after": new_row_count,
            "rows_affected": new_row_count,
            "output_file": output_path,
            "is_read_only": False,
            "changes": [operation_desc, f"Rows: {original_row_count} → {new_row_count}"],
            "data_preview": self._make_json_safe(result_df.head(10).to_dict('records')),
            "generated_code": pandas_code
        }
    
    def _process_with_pattern_matching(self, file_path: str, query: str, output_path: Optional[str] = None, sheet_name: Optional[str] = None) -> Dict:
        """
        Legacy approach: Process query using pattern matching (original implementation)
        Kept for backward compatibility and as fallback.
        """
        # Set output path
        if not output_path:
            base, ext = os.path.splitext(file_path)
            output_path = f"{base}_modified{ext}"
        
        # STEP 1: Analyze Excel Structure
        print("\n📊 STEP 1: Analyzing Excel Structure...")
        excel_structure = self.analyze_excel_structure(file_path, sheet_name=sheet_name)
        
        if "error" in excel_structure:
            return {"status": "failed", "error": excel_structure["error"]}
        
        # Safe column display
        columns_display = ', '.join([str(c) for c in excel_structure.get('columns', [])])
        print(f"   ✓ Found {excel_structure['total_columns']} columns: {columns_display}")
        print(f"   ✓ Total rows: {excel_structure['total_rows']}")
        print(f"   ✓ Sheet: {excel_structure.get('analyzed_sheet', 'N/A')}")
        
        # STEP 2: Parse User Query
        print(f"\n🧠 STEP 2: Analyzing User Intent...")
        print(f"   Query: '{query}'")
        
        parsed_query = self.parse_user_query(query, excel_structure)
        
        if "error" in parsed_query:
            return {"status": "failed", "error": parsed_query["error"]}
        
        # Display reasoning if available
        if "reasoning" in parsed_query:
            reasoning = parsed_query["reasoning"]
            print(f"\n   💭 AI Reasoning:")
            print(f"      Intent: {reasoning.get('intent_summary', 'N/A')}")
            print(f"      Rationale: {reasoning.get('operation_rationale', 'N/A')}")
            if reasoning.get('risk_factors'):
                print(f"      Risk Factors: {', '.join(reasoning.get('risk_factors', []))}")
        
        print(f"   ✓ Operation Type: {parsed_query.get('operation_type')}")
        print(f"   ✓ Sub-Operation: {parsed_query.get('sub_operation')}")
        print(f"   ✓ Source Columns: {', '.join(parsed_query.get('source_columns', []))}")
        print(f"   ✓ Execution Confidence: {parsed_query.get('execution_confidence', 'N/A').upper()}")
        
        # STEP 3: Check for clarification needed
        if parsed_query.get("is_ambiguous") or parsed_query.get("clarification_needed"):
            alternatives = parsed_query.get("alternative_interpretations", [])
            if alternatives:
                print(f"\n   🤔 Alternative Interpretations Detected:")
                for i, alt in enumerate(alternatives, 1):
                    print(f"      {i}. {alt}")
            
            return {
                "status": "clarification_needed",
                "question": parsed_query.get("clarification_needed"),
                "alternatives": alternatives,
                "parsed_info": parsed_query
            }
        
        # STEP 4: Create Decision Plan
        print(f"\n📋 STEP 3: Creating Enterprise Decision Plan...")
        decision = self.create_decision_plan(parsed_query, excel_structure)
        
        print(f"   ✓ Operation: {decision.operation_type} → {decision.sub_operation}")
        print(f"   ✓ Target Column(s): {decision.target_columns if decision.target_columns else 'N/A (visualization/reporting)'}")
        print(f"   ✓ Change Description: {decision.change_description}")
        print(f"   ✓ Risk Level: {decision.risk_level.upper()}")
        
        print("\n   📄 Decision Plan (JSON):")
        print(json.dumps(decision.to_dict(), indent=4))
        
        # STEP 4: Validate Decision Plan
        print(f"\n🔍 STEP 4: Validating Decision Plan...")
        validation_result = self._validate_decision_plan(decision, excel_structure)
        
        if "error" not in validation_result:
            validation_status = validation_result.get("validation_status", "unknown")
            is_valid = validation_result.get("is_valid", False)
            
            print(f"   ✓ Validation Status: {validation_status.upper()}")
            
            # Display warnings if any
            warnings = validation_result.get("warnings", [])
            if warnings:
                print(f"   ⚠️  Warnings:")
                for warning in warnings:
                    print(f"      - {warning}")
            
            # Display blockers if any
            blockers = validation_result.get("blockers", [])
            if blockers:
                print(f"   ❌ Blockers Found:")
                for blocker in blockers:
                    print(f"      - {blocker}")
                return {
                    "status": "validation_failed",
                    "blockers": blockers,
                    "validation_result": validation_result
                }
            
            # Check if user confirmation needed
            if validation_result.get("requires_user_confirmation") or decision.risk_level == "high":
                print(f"   ⚠️  HIGH-RISK OPERATION: User confirmation required")
                return {
                    "status": "confirmation_required",
                    "decision": decision.to_dict(),
                    "validation": validation_result,
                    "message": "This is a high-risk operation. Please review and confirm."
                }
        else:
            print(f"   ⚠️  Validation check skipped (validation service unavailable)")
        
        # STEP 5: Execute Operation
        print(f"\n⚙️  STEP 5: Executing Enterprise Excel Operation...")
        
        try:
            result = self.execute_excel_operation(file_path, decision, output_path)
            
            if result["status"] == "success":
                print(f"   ✅ SUCCESS!")
                print(f"   ✓ Operation: {result.get('operation', 'N/A')}")
                print(f"   ✓ {result['message']}")
                print(f"   ✓ Rows affected: {result['rows_affected']}")
                print(f"   ✓ Output saved to: {result['output_file']}")
                print(f"\n   📋 Changes Made:")
                for change in result.get('changes', []):
                    print(f"      • {change}")
                
                # STEP 6: Generate Business-Friendly Explanation
                print(f"\n📊 STEP 6: Generating Executive Summary...")
                explanation = self._generate_result_explanation(result, decision)
                
                if "error" not in explanation:
                    print(f"\n   Executive Summary:")
                    print(f"   {explanation.get('executive_summary', 'Operation completed successfully')}")
                    
                    # Add explanation to result
                    result['explanation'] = explanation
                
                # Log the change for audit
                self.change_log.append({
                    "timestamp": pd.Timestamp.now().isoformat(),
                    "file": file_path,
                    "query": query,
                    "operation": result.get('operation'),
                    "output": output_path,
                    "risk_level": decision.risk_level,
                    "decision": decision.to_dict()
                })
            else:
                print(f"   ❌ FAILED: {result.get('error')}")
                print(f"   Operation attempted: {result.get('operation', 'N/A')}")
        
        except Exception as e:
            error_msg = str(e)
            print(f"   ❌ EXECUTION ERROR: {error_msg}")
            
            # STEP 6 (Error Path): Generate Recovery Suggestions
            print(f"\n🔧 STEP 6: Analyzing Error and Generating Recovery Strategies...")
            recovery = self._handle_error_recovery(
                error=error_msg,
                query=query,
                context={
                    "excel_structure": excel_structure,
                    "decision": decision.to_dict(),
                    "parsed_query": parsed_query
                }
            )
            
            if "error" not in recovery:
                print(f"\n   Error Analysis:")
                print(f"   Category: {recovery.get('error_category', 'unknown')}")
                print(f"   Root Cause: {recovery.get('root_cause', 'Unknown')}")
                print(f"   User Message: {recovery.get('user_friendly_message', '')}")
                
                suggested_fixes = recovery.get('suggested_fixes', [])
                if suggested_fixes:
                    print(f"\n   💡 Suggested Fixes:")
                    for i, fix in enumerate(suggested_fixes, 1):
                        print(f"      {i}. {fix.get('fix_description')}")
                        print(f"         Try: \"{fix.get('modified_query')}\"")
                        print(f"         Success Likelihood: {fix.get('success_likelihood', 'unknown')}")
            
            result = {
                "status": "failed",
                "error": error_msg,
                "recovery_suggestions": recovery
            }
        
        print("\n" + "="*60)
        
        return result
    
    def get_audit_log(self) -> List[Dict]:
        """
        Get complete audit trail of all operations performed by the agent
        
        Returns:
            List of audit log entries with timestamps, queries, and operations
        """
        return self.change_log
    
    def export_audit_log(self, output_path: str = "agent_audit_log.json") -> str:
        """
        Export audit log to JSON file for compliance and review
        
        Args:
            output_path: Path to save the audit log
            
        Returns:
            Path to the exported file
        """
        with open(output_path, 'w') as f:
            json.dump(self.change_log, f, indent=2, default=str)
        
        return output_path


# ==========================================
# HELPER FUNCTION FOR QUICK USAGE
# ==========================================
def create_excel_agent(api_key: str) -> ExcelAgent:
    """Factory function to create an Excel Agent instance"""
    return ExcelAgent(api_key=api_key)


# ==========================================
# INTERACTIVE MODE - MAIN ENTRY POINT
# ==========================================
def run_interactive_agent(api_key: str):
    """
    Interactive mode where agent asks for Excel file and then processes queries
    """
    print("\n" + "="*70)
    print("🏢 ENTERPRISE EXCEL AUTOMATION AI AGENT - INTERACTIVE MODE")
    print("="*70)
    print("\nWelcome! I'm your Enterprise Excel Automation Agent.")
    print("I can help you with professional Excel operations using business language.\n")
    print("I am NOT a chatbot - I am a decision-making automation expert.")
    print("\n📊 SUPPORTED OPERATIONS:")
    print("  1. Calculations: Sum, Average, Min, Max, %, Count, Median")
    print("  2. Transformations: Add/Remove/Rename columns, Sort, Filter, Deduplicate")
    print("  3. Validations: Dropdowns, Error detection, Quality checks")
    print("  4. Formatting: Conditional colors, Highlights, Data bars")
    print("  5. Aggregation: Pivot tables, Group by, Summaries")
    print("  6. Lookups: XLOOKUP, VLOOKUP, Data matching")
    print("  7. Visualization: Charts (Bar, Pie, Line), KPI dashboards")
    print("  8. Reporting: Summary sheets, Change logs\n")
    
    # Initialize agent
    agent = ExcelAgent(api_key=api_key)
    
    # STEP 1: Ask for Excel file
    print("="*70)
    print("STEP 1: Excel File Selection")
    print("="*70)
    
    while True:
        file_path = input("\n📁 Please provide the path to your Excel file: ").strip()
        
        # Remove quotes if user wrapped path in quotes
        file_path = file_path.strip('"').strip("'")
        
        if not file_path:
            print("⚠️  No file path provided. Please try again.")
            continue
        
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            retry = input("Would you like to try again? (yes/no): ").strip().lower()
            if retry not in ['yes', 'y']:
                print("\n👋 Exiting. Goodbye!")
                return
            continue
        
        # Try to load and analyze the file
        try:
            print(f"\n📊 Analyzing Excel file: {os.path.basename(file_path)}")
            structure = agent.analyze_excel_structure(file_path)
            
            if "error" in structure:
                print(f"❌ Error reading file: {structure['error']}")
                retry = input("Would you like to try another file? (yes/no): ").strip().lower()
                if retry not in ['yes', 'y']:
                    print("\n👋 Exiting. Goodbye!")
                    return
                continue
            
            # Display file information
            print(f"\n✅ Successfully loaded: {structure['file_name']}")
            
            # Handle multiple sheets - let user select
            available_sheets = structure.get('available_sheets', [])
            selected_sheet = structure['analyzed_sheet']
            
            if len(available_sheets) > 1:
                print(f"\n📑 Multiple sheets detected ({len(available_sheets)} sheets):")
                for idx, sheet in enumerate(available_sheets, 1):
                    marker = "→" if sheet == selected_sheet else " "
                    print(f"   {marker} {idx}. {sheet}")
                
                sheet_choice = input("\n📄 Enter sheet number or name to work with (default: 1): ").strip()
                
                if sheet_choice:
                    # Try to parse as number
                    try:
                        sheet_idx = int(sheet_choice) - 1
                        if 0 <= sheet_idx < len(available_sheets):
                            selected_sheet = available_sheets[sheet_idx]
                        else:
                            print(f"⚠️  Invalid sheet number. Using first sheet.")
                            selected_sheet = available_sheets[0]
                    except ValueError:
                        # Treat as sheet name
                        if sheet_choice in available_sheets:
                            selected_sheet = sheet_choice
                        else:
                            print(f"⚠️  Sheet '{sheet_choice}' not found. Using first sheet.")
                            selected_sheet = available_sheets[0]
                
                # Re-analyze with selected sheet if different
                if selected_sheet != structure['analyzed_sheet']:
                    structure = agent.analyze_excel_structure(file_path, sheet_name=selected_sheet)
            
            print(f"\n   📄 Working with sheet: {structure['analyzed_sheet']}")
            print(f"   📊 Rows: {structure['total_rows']}")
            print(f"   📋 Columns ({structure['total_columns']}): {', '.join([str(c) for c in structure['columns']])}")
            
            # Show header warning if present
            if structure.get('header_warning'):
                print(f"\n   {structure['header_warning']}")
            
            # Show sample data
            if structure.get('sample_data'):
                print(f"\n   Sample Data (first 3 rows):")
                sample_df = pd.DataFrame(structure['sample_data'])
                print("   " + "\n   ".join(sample_df.to_string(index=False).split('\n')[:6]))
            
            break
            
        except Exception as e:
            print(f"❌ Error: {e}")
            retry = input("Would you like to try another file? (yes/no): ").strip().lower()
            if retry not in ['yes', 'y']:
                print("\n👋 Exiting. Goodbye!")
                return
            continue
    
    # STEP 2: Interactive query loop
    print("\n" + "="*70)
    print("STEP 2: Modification Queries")
    print("="*70)
    print("\nYou can now ask me to perform operations using business language.")
    print("\n💼 EXAMPLE BUSINESS QUERIES:")
    print("  📊 Calculations:")
    print("    • 'Calculate total sales by summing Q1, Q2, and Q3'")
    print("    • 'Find the average revenue across all products'")
    print("    • 'Calculate profit margin percentage'")
    print("\n  📈 Visualizations:")
    print("    • 'Create a bar chart for sales by product'")
    print("    • 'Show me a pie chart of revenue distribution'")
    print("\n  🔧 Transformations:")
    print("    • 'Sort data by Sales descending'")
    print("    • 'Remove duplicate entries'")
    print("    • 'Replace NaN values with N/A'")
    print("\n  🎨 Formatting:")
    print("    • 'Highlight cells where profit is negative'")
    print("\n📌 COMMANDS:")
    print("  • Type 'quit', 'exit', or 'q' to stop.")
    print("  • Type 'new' to load a different Excel file.")
    print("  • Type 'sheets' to list available sheets.")
    print("  • Type 'switch sheet <name>' to switch to a different sheet.")
    print("="*70)
    
    query_count = 0
    current_file = file_path
    current_sheet = selected_sheet
    current_structure = structure

    while True:
        print(f"\n{'─'*70}")
        print(f"📄 Current: {os.path.basename(current_file)} → Sheet: {current_sheet}")
        query = input("\n💬 What would you like to modify in the Excel? ").strip()
        
        # Handle exit commands
        if query.lower() in ['quit', 'exit', 'q']:
            print("\n" + "="*70)
            print("👋 Thank you for using Excel Automation Agent!")
            print("="*70)
            break
        
        # Handle new file command
        if query.lower() == 'new':
            print("\n🔄 Loading a new Excel file...\n")
            run_interactive_agent(api_key)
            return
        
        # Handle sheets listing command
        if query.lower() == 'sheets':
            print(f"\n📑 Available sheets in {os.path.basename(current_file)}:")
            for idx, sheet in enumerate(current_structure.get('available_sheets', []), 1):
                marker = "→" if sheet == current_sheet else " "
                print(f"   {marker} {idx}. {sheet}")
            continue
        
        # Handle switch sheet command
        if query.lower().startswith('switch sheet'):
            sheet_name = query[12:].strip()
            if not sheet_name:
                print("⚠️  Please specify a sheet name. Usage: switch sheet <name>")
                continue
            
            available_sheets = current_structure.get('available_sheets', [])
            
            # Try to find the sheet (by name or number)
            target_sheet = None
            try:
                sheet_idx = int(sheet_name) - 1
                if 0 <= sheet_idx < len(available_sheets):
                    target_sheet = available_sheets[sheet_idx]
            except ValueError:
                if sheet_name in available_sheets:
                    target_sheet = sheet_name
            
            if target_sheet:
                print(f"\n🔄 Switching to sheet: {target_sheet}")
                current_structure = agent.analyze_excel_structure(current_file, sheet_name=target_sheet)
                current_sheet = target_sheet
                print(f"   📊 Rows: {current_structure['total_rows']}")
                print(f"   📋 Columns: {', '.join([str(c) for c in current_structure['columns']])}")
            else:
                print(f"❌ Sheet '{sheet_name}' not found.")
                print(f"   Available sheets: {', '.join(available_sheets)}")
            continue
        
        # Skip empty queries
        if not query:
            print("⚠️  Please enter a query or type 'quit' to exit.")
            continue
        
        # Process the query
        query_count += 1
        
        # Generate output filename
        base_name = os.path.splitext(os.path.basename(current_file))[0]
        output_path = f"{base_name}_modified_{query_count}.xlsx"
        
        print(f"\n🔄 Processing your request...")
        
        try:
            result = agent.process_query(
                file_path=current_file,
                query=query,
                output_path=output_path,
                sheet_name=current_sheet
            )
            
            # Handle results
            if result.get("status") == "success":
                # Check if this was a read-only operation
                if result.get("is_read_only"):
                    print(f"\n{'='*70}")
                    print("📊 DATA DISPLAY (Read-Only)")
                    print(f"{'='*70}")
                    print(f"📝 {result['message']}")
                    print(f"📊 Total rows: {result['rows_affected']}")
                    
                    # Show data preview
                    if result.get('data_preview'):
                        print(f"\n📋 Data Preview (first 10 rows):")
                        preview_df = pd.DataFrame(result['data_preview'])
                        print(preview_df.to_string(index=False))
                    
                    # No file was modified for read-only operations
                    query_count -= 1  # Don't count read-only queries
                else:
                    print(f"\n{'='*70}")
                    print("✅ MODIFICATION SUCCESSFUL!")
                    print(f"{'='*70}")
                    print(f"📁 New file created: {result['output_file']}")
                    print(f"📝 {result['message']}")
                    print(f"📊 Rows modified: {result['rows_affected']}")
                    
                    # Show preview
                    try:
                        preview_df = pd.read_excel(output_path, sheet_name=current_sheet)
                        print(f"\n📋 Preview of modified data (first 10 rows):")
                        print(preview_df.head(10).to_string(index=False))
                    except:
                        pass
                    
                    # Ask if they want to continue with the modified file
                    print(f"\n{'─'*70}")
                    continue_choice = input("\n🔄 Use this modified file for next query? (yes/no, default: yes): ").strip().lower()
                    
                    if continue_choice in ['', 'yes', 'y']:
                        current_file = output_path
                        # Re-analyze the structure for the new file
                        current_structure = agent.analyze_excel_structure(current_file, sheet_name=current_sheet)
                        print(f"✅ Now working with: {output_path}")
                    else:
                        print(f"✅ Still working with: {current_file}")
            
            elif result.get("status") == "clarification_needed":
                print(f"\n❓ I need clarification:")
                print(f"   {result.get('question', 'Please provide more details.')}")
                print("\n💡 Please rephrase your query with more details.")
                query_count -= 1  # Don't count failed queries
            
            else:
                print(f"\n❌ Error: {result.get('error', 'Unknown error occurred')}")
                print("💡 Please try rephrasing your query or check the column names.")
                query_count -= 1  # Don't count failed queries
        
        except Exception as e:
            print(f"\n❌ Unexpected error: {e}")
            print("💡 Please try again with a different query.")
            query_count -= 1  # Don't count failed queries


# ==========================================
# MAIN ENTRY POINT
# ==========================================
if __name__ == "__main__":
    import os
    
    # Get API Key from environment variable or user input
    API_KEY = os.getenv("GROQ_API_KEY")
    
    if not API_KEY:
        print("⚠️  GROQ_API_KEY not found in environment variables")
        print("Get your free API key from: https://console.groq.com\n")
        API_KEY = input("Enter your Groq API Key: ").strip()
    
    # Run interactive agent
    run_interactive_agent(API_KEY)
